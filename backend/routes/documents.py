from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from datetime import datetime, timezone, timedelta
from typing import List, Optional
import uuid
import os
import logging
import tempfile
import aiofiles
import re
import openpyxl
import xlsxwriter
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

from database import db
from models import (
    TransactionStatus, SourceType, SRI_CATEGORIES, SUBSCRIPTION_SERVICES,
    INTERNATIONAL_COUNTRIES, STATUS_LABELS, CANASTA_BASICA, FRACCION_BASICA_EXENTA,
    CARGAS_FAMILIARES_CBF, PORCENTAJE_REBAJA_IR, CONTRIBUYENTE_INFO, UserRole
)
from utils import (
    get_current_user, check_role, classify_with_ai, process_image_with_ai,
    lookup_known_vendor, find_potential_duplicates, dedup_or_merge, UPLOADS_DIR
)

logger = logging.getLogger(__name__)
router = APIRouter()


@router.post("/process/email")
async def process_email(email_content: str = Form(...), user: dict = Depends(get_current_user)):
    result = await classify_with_ai(email_content, "email")
    transaction_id = str(uuid.uuid4())
    doc = {"id": transaction_id, "user_id": user["id"], "amount": result.get("amount", 0), "description": result.get("description", ""), "category": result.get("category", "otros"), "subcategory": result.get("subcategory", "Varios"), "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"), "transaction_type": "expense", "establishment": result.get("establishment", ""), "ai_classified": True, "created_at": datetime.now(timezone.utc).isoformat()}
    await db.transactions.insert_one(doc)
    return {"message": "Email procesado", "transaction": {k: v for k, v in doc.items() if k != "_id"}}


@router.post("/process/receipt")
async def process_receipt(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    temp_dir = tempfile.mkdtemp()
    file_path = os.path.join(temp_dir, file.filename)
    async with aiofiles.open(file_path, 'wb') as f:
        content = await file.read()
        await f.write(content)
    try:
        result = await process_image_with_ai(file_path)
        created_transactions = []
        for t in result.get("transactions", []):
            transaction_id = str(uuid.uuid4())
            doc = {"id": transaction_id, "user_id": user["id"], "amount": t.get("amount", 0), "description": t.get("description", ""), "category": t.get("category", "otros"), "subcategory": t.get("subcategory", "Varios"), "date": t.get("date", datetime.now(timezone.utc).strftime("%Y-%m-%d")), "transaction_type": "expense", "establishment": t.get("establishment", ""), "ai_classified": True, "created_at": datetime.now(timezone.utc).isoformat()}
            await db.transactions.insert_one(doc)
            created_transactions.append({k: v for k, v in doc.items() if k != "_id"})
        return {"message": f"Procesadas {len(created_transactions)} transacciones", "transactions": created_transactions}
    finally:
        if os.path.exists(file_path):
            os.remove(file_path)
        os.rmdir(temp_dir)


@router.post("/process/receipts-multiple")
async def process_multiple_receipts(files: List[UploadFile] = File(...), user: dict = Depends(get_current_user)):
    all_transactions = []
    errors = []
    for file in files:
        temp_dir = tempfile.mkdtemp()
        file_path = os.path.join(temp_dir, file.filename)
        try:
            async with aiofiles.open(file_path, 'wb') as f:
                content = await file.read()
                await f.write(content)
            result = await process_image_with_ai(file_path)
            for t in result.get("transactions", []):
                country = t.get("country", "Ecuador")
                is_international = any(c.lower() in country.lower() for c in INTERNATIONAL_COUNTRIES) if country else False
                amount = t.get("amount", 0)
                date = t.get("date", datetime.now(timezone.utc).strftime("%Y-%m-%d"))
                establishment = t.get("establishment", "")
                description = t.get("description", "")
                vendor_lookup = await lookup_known_vendor(user["id"], establishment, description)
                if vendor_lookup["found"]:
                    category = vendor_lookup.get("personal_category") or t.get("category", "otros")
                    sri_category = vendor_lookup.get("sri_category")
                    subcategory = vendor_lookup.get("subcategory") or t.get("subcategory", "Varios")
                    is_deductible = vendor_lookup.get("is_deductible", False)
                    auto_categorized_by = "known_vendor"
                else:
                    category = "viajes_internacionales" if is_international else t.get("category", "otros")
                    sri_category = t.get("sri_category")
                    subcategory = t.get("subcategory", "Varios")
                    is_deductible = SRI_CATEGORIES.get(category, {}).get("deductible", False)
                    auto_categorized_by = "ai" if t.get("category") else None
                duplicates = await find_potential_duplicates(user["id"], amount, date, establishment, description)
                transaction_id = str(uuid.uuid4())
                doc = {"id": transaction_id, "user_id": user["id"], "amount": amount, "description": description, "category": category, "personal_category": category, "sri_category": sri_category, "subcategory": subcategory, "date": date, "transaction_type": "expense", "establishment": establishment, "country": country, "is_international": is_international, "payment_source": "internacional" if is_international else "local", "is_deductible": is_deductible, "ai_classified": True, "auto_categorized_by": auto_categorized_by, "status": TransactionStatus.DUPLICATE_SUSPECT if duplicates else TransactionStatus.PENDING_REVIEW, "source_type": SourceType.RECEIPT, "has_receipt": True, "duplicate_of": duplicates[0]["transaction"]["id"] if duplicates else None, "match_confidence": duplicates[0]["confidence"] if duplicates else None, "created_at": datetime.now(timezone.utc).isoformat(), "source_file": file.filename}
                await db.transactions.insert_one(doc)
                all_transactions.append({k: v for k, v in doc.items() if k != "_id"})
                try:
                    from routes.sri_match import try_sri_match, retry_pending_matches
                    await try_sri_match(user["id"], transaction_id)
                    await retry_pending_matches(user["id"])
                except Exception as e:
                    logger.warning(f"SRI match hook failed: {e}")
        except Exception as e:
            errors.append({"file": file.filename, "error": str(e)})
        finally:
            if os.path.exists(file_path):
                os.remove(file_path)
            if os.path.exists(temp_dir):
                os.rmdir(temp_dir)
    return {"message": f"Procesados {len(files)} archivos, {len(all_transactions)} transacciones creadas", "transactions": all_transactions, "errors": errors}


def _regex_fill_card_info(card_info: dict, raw_text: str) -> dict:
    """Complementa card_info con campos faltantes usando regex sobre el texto crudo.
    Solo rellena campos que están None/missing. No sobreescribe valores ya presentes."""
    if not raw_text:
        return card_info

    def _parse_num(s: str) -> Optional[float]:
        s = (s or "").replace(" ", "").replace(",", "")
        # ES "1.234,56" -> "1234.56" ya quedó arriba sin coma; handle dot-thousands
        try:
            return float(s)
        except ValueError:
            return None

    def _find_first(patterns, text):
        for pat in patterns:
            m = re.search(pat, text, re.IGNORECASE)
            if m:
                val = _parse_num(m.group(1))
                if val is not None and val > 0:
                    return val
        return None

    if card_info.get("credit_limit") is None:
        card_info["credit_limit"] = _find_first([
            r"CUPO\s*AUTORIZADO[^\d]{0,30}\$?\s*([\d.,]+)",
            r"Cupo\s*Aprobado[^\d]{0,30}\$?\s*([\d.,]+)",
            r"L[IÍ]MITE\s*DE\s*CR[EÉ]DITO[^\d]{0,30}\$?\s*([\d.,]+)",
        ], raw_text)

    if card_info.get("pago_total") is None:
        card_info["pago_total"] = _find_first([
            r"PAGO\s*DE\s*CONTADO[^\d]{0,30}\$?\s*([\d.,]+)",
            r"TOTAL\s*A\s*PAGAR[^\d]{0,30}\$?\s*([\d.,]+)",
            r"Total\s*a\s*pagar[^\d]{0,30}\$?\s*([\d.,]+)",
        ], raw_text)

    if card_info.get("deferred_balance") is None:
        card_info["deferred_balance"] = _find_first([
            r"SALDO\s*ACTUAL\s*DIFERIDO[^\d]{0,30}\$?\s*([\d.,]+)",
            r"SALDO\s*DIFERIDO[^\d]{0,30}\$?\s*([\d.,]+)",
            r"Cr[eé]dito\s*Diferido[^\d]{0,30}\$?\s*([\d.,]+)",
        ], raw_text)

    # Fix due_date year if it appears in the past vs statement_date: assume same or next year
    due = card_info.get("due_date")
    stmt = card_info.get("statement_date")
    if due and stmt:
        try:
            due_dt = datetime.strptime(due[:10], "%Y-%m-%d")
            stmt_dt = datetime.strptime(stmt[:10], "%Y-%m-%d")
            if due_dt < stmt_dt:
                # due_date debe ser >= statement_date; subir al mismo año o siguiente
                new_due = due_dt.replace(year=stmt_dt.year)
                if new_due < stmt_dt:
                    new_due = new_due.replace(year=stmt_dt.year + 1)
                card_info["due_date"] = new_due.strftime("%Y-%m-%d")
        except Exception:
            pass
    return card_info


async def _upsert_card_from_statement(user_id: str, card_info: dict, response_data: dict, raw_text: str = ""):
    """Extract or update credit card info from statement."""
    card_info = _regex_fill_card_info(dict(card_info or {}), raw_text)
    bank_name = card_info.get("bank_name", "").lower()
    card_number = card_info.get("card_number_last4", "")
    existing_card = None
    if bank_name:
        # Matching case-insensitive + normalización: sin espacios, lowercase.
        # "PacifiCard", "PACIFICARD", "Banco del Pacífico PacifiCard" matchean
        # al mismo registro si uno contiene al otro.
        def _norm(s: str) -> str:
            return "".join((s or "").lower().split())
        norm_bank = _norm(bank_name)
        cards = await db.credit_cards.find({"user_id": user_id}).to_list(200)
        for c in cards:
            if card_number and c.get("last_four_digits") == card_number:
                existing_card = c
                break
            norm_name = _norm(c.get("name", ""))
            norm_c_bank = _norm(c.get("bank", ""))
            if norm_bank and (
                norm_bank in norm_name or norm_name in norm_bank
                or (norm_c_bank and (norm_bank in norm_c_bank or norm_c_bank in norm_bank))
            ):
                existing_card = c
                break

    if existing_card:
        update_data = {"current_balance": card_info.get("current_balance", existing_card.get("current_balance", 0)), "minimum_payment": card_info.get("minimum_payment", existing_card.get("minimum_payment", 0)), "credit_limit": card_info.get("credit_limit") if card_info.get("credit_limit") is not None else existing_card.get("credit_limit", 0), "available_credit": card_info.get("available_credit"), "statement_date": card_info.get("statement_date"), "due_date": card_info.get("due_date"), "saldo_diferido": card_info.get("deferred_balance") if card_info.get("deferred_balance") is not None else existing_card.get("saldo_diferido"), "pago_total": card_info.get("pago_total") if card_info.get("pago_total") is not None else existing_card.get("pago_total"), "updated_at": datetime.now(timezone.utc).isoformat()}
        if card_info.get("apr"):
            update_data["apr"] = card_info["apr"]
        await db.credit_cards.update_one({"id": existing_card["id"]}, {"$set": update_data})
        response_data["card_updated"] = True
        response_data["card_info"] = {**existing_card, **update_data, "_id": None}
        card_id = existing_card["id"]
        card_name_val = existing_card["name"]
    else:
        card_id = str(uuid.uuid4())
        new_card = {"id": card_id, "user_id": user_id, "name": card_info.get("card_name") or card_info.get("bank_name", "Tarjeta Importada"), "bank": card_info.get("bank_name", ""), "last_four_digits": card_number, "credit_limit": card_info.get("credit_limit", 0), "current_balance": card_info.get("current_balance", 0), "minimum_payment": card_info.get("minimum_payment", 0), "apr": card_info.get("apr", 0), "statement_date": card_info.get("statement_date"), "due_date": card_info.get("due_date"), "available_credit": card_info.get("available_credit"), "saldo_diferido": card_info.get("deferred_balance"), "pago_total": card_info.get("pago_total"), "currency": "USD", "is_international": False, "created_at": datetime.now(timezone.utc).isoformat()}
        await db.credit_cards.insert_one(new_card)
        response_data["card_info"] = {k: v for k, v in new_card.items() if k != "_id"}
        response_data["card_updated"] = True
        card_name_val = new_card["name"]

    # Schedule payment if due date known
    if card_info.get("due_date") and card_info.get("current_balance"):
        due_date = card_info["due_date"]
        try:
            due_day = int(due_date.split("-")[2]) if "-" in due_date else 15
        except Exception:
            due_day = 15
        existing_payment = await db.scheduled_payments.find_one({"user_id": user_id, "card_id": card_id, "month": datetime.now().month})
        if not existing_payment:
            payment_doc = {"id": str(uuid.uuid4()), "user_id": user_id, "card_id": card_id, "description": f"Pago Tarjeta {card_name_val}", "amount": card_info.get("current_balance", 0), "minimum_amount": card_info.get("minimum_payment", 0), "due_day": due_day, "due_date": due_date, "month": datetime.now().month, "year": datetime.now().year, "payment_method": "transferencia", "category": "tarjeta_credito", "is_recurring": True, "is_card_payment": True, "status": "pending", "created_at": datetime.now(timezone.utc).isoformat()}
            await db.scheduled_payments.insert_one(payment_doc)
            response_data["payment_scheduled"] = True


async def _save_deferred_purchases(user_id: str, deferred_purchases: list, card_info: dict, response_data: dict, filename: str):
    """Save deferred purchase records. If a deferred with same description+card already exists
    as active, decrement its remaining_installments instead of creating a duplicate."""
    deferred_created = []
    deferred_decremented = []
    card_name_val = (card_info.get("card_name") or card_info.get("bank_name")) if card_info else None
    for dp in deferred_purchases:
        if dp.get("remaining_installments", 0) <= 0:
            continue
        description = dp.get("description", "Compra Diferida")
        # Look up an existing active deferred with the same description + card
        existing = await db.deferred_payments.find_one({
            "user_id": user_id,
            "description": {"$regex": f"^{re.escape(description)}$", "$options": "i"},
            "card_name": card_name_val,
            "remaining_installments": {"$gt": 0},
            "total_amount": {
                "$gte": dp.get("total_amount", 0) * 0.95,
                "$lte": dp.get("total_amount", 0) * 1.05
            }
        })
        if existing:
            remaining = existing.get("remaining_installments", 0)
            monthly_payment = existing.get("monthly_payment", dp.get("monthly_payment", 0))
            new_remaining = max(0, remaining - 1)
            new_remaining_amount = new_remaining * monthly_payment
            await db.deferred_payments.update_one(
                {"id": existing["id"]},
                {"$set": {
                    "remaining_installments": new_remaining,
                    "remaining_amount": new_remaining_amount,
                    "last_payment_date": datetime.now(timezone.utc).isoformat(),
                    "paid_installments": existing.get("paid_installments", 0) + 1,
                }}
            )
            deferred_decremented.append({"id": existing["id"], "description": description, "remaining_installments": new_remaining})
        else:
            deferred_doc = {"id": str(uuid.uuid4()), "user_id": user_id, "description": description, "total_amount": dp.get("total_amount", 0), "monthly_payment": dp.get("monthly_payment", 0), "remaining_installments": dp.get("remaining_installments", 0), "total_installments": dp.get("total_installments", dp.get("remaining_installments", 0)), "card_id": response_data.get("card_info", {}).get("id") if response_data.get("card_info") else None, "card_name": card_name_val, "created_at": datetime.now(timezone.utc).isoformat(), "source_file": filename}
            await db.deferred_payments.insert_one(deferred_doc)
            deferred_created.append({k: v for k, v in deferred_doc.items() if k != "_id"})
    response_data["deferred_payments_created"] = len(deferred_created)
    response_data["deferred_payments_decremented"] = len(deferred_decremented)
    response_data["deferred_payments"] = deferred_created
    return deferred_created


def _categorize_transaction(t: dict, vendor_lookup: dict, is_subscription: bool) -> dict:
    """Determine category, sri_category, subcategory, deductibility for a transaction."""
    if vendor_lookup["found"]:
        return {
            "category": vendor_lookup.get("personal_category") or t.get("category", "otros"),
            "sri_category": vendor_lookup.get("sri_category"),
            "subcategory": vendor_lookup.get("subcategory") or t.get("subcategory", "Varios"),
            "is_deductible": vendor_lookup.get("is_deductible", False),
            "auto_categorized_by": "known_vendor"
        }
    if is_subscription:
        return {
            "category": "suscripciones", "sri_category": None,
            "subcategory": t.get("subcategory", "Varios"),
            "is_deductible": False, "auto_categorized_by": "subscription_detection"
        }
    category = t.get("category", "otros")
    return {
        "category": category, "sri_category": t.get("sri_category"),
        "subcategory": t.get("subcategory", "Varios"),
        "is_deductible": SRI_CATEGORIES.get(category, {}).get("deductible", False),
        "auto_categorized_by": "ai" if t.get("category") else None
    }


async def _save_statement_transactions(user_id: str, transactions: list, card_info: dict, filename: str) -> list:
    """Validate, categorize and save each transaction from a bank statement."""
    created = []
    for t in transactions:
        amount = abs(t.get("amount", 0))
        if amount == 0:
            continue
        is_payment = t.get("amount", 0) < 0 or "pago" in t.get("description", "").lower()
        if is_payment:
            continue

        date = t.get("date", datetime.now(timezone.utc).strftime("%Y-%m-%d"))
        establishment = t.get("establishment", t.get("description", "")[:50])
        description = t.get("description", "")

        vendor_lookup = await lookup_known_vendor(user_id, establishment, description)
        is_subscription = t.get("is_subscription", False)
        text_lower = f"{description} {establishment}".lower()
        for sub in SUBSCRIPTION_SERVICES:
            if sub in text_lower:
                is_subscription = True
                break

        cat_info = _categorize_transaction(t, vendor_lookup, is_subscription)
        duplicates = await find_potential_duplicates(user_id, amount, date, establishment, description)

        doc = {
            "id": str(uuid.uuid4()), "user_id": user_id, "amount": amount,
            "description": description, "category": cat_info["category"],
            "personal_category": cat_info["category"], "sri_category": cat_info["sri_category"],
            "subcategory": cat_info["subcategory"], "date": date,
            "transaction_type": "expense", "establishment": establishment,
            "is_international": t.get("is_international", False) if not is_subscription else False,
            "is_subscription": is_subscription, "is_recurring": is_subscription,
            "is_deductible": cat_info["is_deductible"],
            "tags": ["recurrente", "suscripcion"] if is_subscription else [],
            "payment_method": "tarjeta",
            "card_name": card_info.get("card_name") or card_info.get("bank_name") if card_info else None,
            "ai_classified": True, "auto_categorized_by": cat_info["auto_categorized_by"],
            "status": TransactionStatus.DUPLICATE_SUSPECT if duplicates else TransactionStatus.PENDING_REVIEW,
            "source_type": SourceType.BANK_STATEMENT,
            "duplicate_of": duplicates[0]["transaction"]["id"] if duplicates else None,
            "match_confidence": duplicates[0]["confidence"] if duplicates else None,
            "created_at": datetime.now(timezone.utc).isoformat(), "source_file": filename,
            "tarjeta_ultimos4": card_info.get("card_number_last4") if card_info else None
        }
        result = await dedup_or_merge(user_id, doc, "estado_cuenta")
        if result["action"] == "inserted":
            created.append({k: v for k, v in doc.items() if k != "_id"})
        else:
            created.append({**result["doc"], "_merged": True})
    return created


@router.post("/process/bank-statement")
async def process_bank_statement(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    temp_dir = tempfile.mkdtemp()
    file_path = os.path.join(temp_dir, file.filename)
    try:
        async with aiofiles.open(file_path, 'wb') as f:
            content = await file.read()
            await f.write(content)

        result = await process_image_with_ai(file_path, document_type="bank_statement")
        response_data = {"card_info": None, "card_updated": False, "transactions_created": 0, "transactions": []}
        card_info = result.get("card_info") or {}

        # 1. Upsert credit card
        if card_info.get("current_balance"):
            # Extract raw text from PDF once for regex fallback
            try:
                from utils import extract_text_from_pdf
                raw_text = extract_text_from_pdf(file_path) or ""
            except Exception:
                raw_text = ""
            await _upsert_card_from_statement(user["id"], card_info, response_data, raw_text=raw_text)

        # 2. Save deferred purchases
        deferred_created = await _save_deferred_purchases(
            user["id"], result.get("deferred_purchases", []),
            card_info, response_data, file.filename
        )

        # 3. Process and save transactions
        created_transactions = await _save_statement_transactions(
            user["id"], result.get("transactions", []),
            card_info, file.filename
        )

        response_data["transactions_created"] = len(created_transactions)
        response_data["transactions"] = created_transactions
        return {"message": f"Estado de cuenta procesado: {len(created_transactions)} transacciones, {len(deferred_created)} diferidos", "data": response_data, "raw_extraction": result if not created_transactions else None}
    except Exception as e:
        logger.error(f"Bank statement processing error: {e}")
        raise HTTPException(status_code=500, detail=f"Error procesando estado de cuenta: {str(e)}")
    finally:
        if os.path.exists(file_path):
            os.remove(file_path)
        if os.path.exists(temp_dir):
            os.rmdir(temp_dir)


@router.post("/process/excel")
async def process_excel(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(BytesIO(content))
        budget_items = []
        total_income = 0
        total_expenses = 0
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] and isinstance(row[0], str):
                    category_name = str(row[0]).lower()
                    category_map = {"servicios basicos": ("vivienda", "Servicios basicos"), "empleados": ("otros", "Empleados"), "colegio": ("educacion", "Colegio y actividades"), "seguros": ("salud", "Seguros"), "comida": ("alimentacion", "Comida"), "restaurantes": ("alimentacion", "Restaurantes"), "carros": ("transporte", "Carros"), "viajes": ("otros", "Viajes y Entretenimiento")}
                    for key, (cat, subcat) in category_map.items():
                        if key in category_name:
                            amounts = [v for v in row[1:] if isinstance(v, (int, float))]
                            if amounts:
                                avg_amount = sum(amounts) / len(amounts)
                                budget_items.append({"category": cat, "subcategory": subcat, "planned_amount": avg_amount, "month": datetime.now(timezone.utc).strftime("%Y-%m")})
                                total_expenses += avg_amount
                            break
                    if "ingreso" in category_name or "personal" in category_name or "apx" in category_name:
                        amounts = [v for v in row[1:] if isinstance(v, (int, float))]
                        if amounts:
                            total_income += sum(amounts) / len(amounts)
        budget_id = str(uuid.uuid4())
        budget_doc = {"id": budget_id, "user_id": user["id"], "items": budget_items, "total_income": total_income, "total_expenses": total_expenses, "created_at": datetime.now(timezone.utc).isoformat()}
        await db.budgets.insert_one(budget_doc)
        return {"message": "Excel procesado", "budget": {k: v for k, v in budget_doc.items() if k != "_id"}}
    except Exception as e:
        logger.error(f"Excel processing error: {e}")
        raise HTTPException(status_code=400, detail=f"Error procesando Excel: {str(e)}")


# ================= ATTACHMENTS =================

@router.post("/transactions/{transaction_id}/attachments")
async def upload_attachment(transaction_id: str, file: UploadFile = File(...), attachment_type: str = Form("receipt"), user: dict = Depends(get_current_user)):
    transaction = await db.transactions.find_one({"id": transaction_id, "user_id": user["id"]}, {"_id": 0})
    if not transaction:
        raise HTTPException(status_code=404, detail="Transaccion no encontrada")
    file_ext = file.filename.split(".")[-1] if "." in file.filename else "jpg"
    filename = f"{transaction_id}_{attachment_type}_{uuid.uuid4().hex[:8]}.{file_ext}"
    file_path = UPLOADS_DIR / filename
    async with aiofiles.open(file_path, 'wb') as f:
        content = await file.read()
        await f.write(content)
    attachments = transaction.get("attachments", [])
    attachments.append(str(filename))
    update_data = {"attachments": attachments}
    if attachment_type == "receipt":
        update_data["has_receipt"] = True
    elif attachment_type == "invoice":
        update_data["has_invoice"] = True
    await db.transactions.update_one({"id": transaction_id}, {"$set": update_data})
    return {"message": "Archivo adjuntado", "filename": filename, "type": attachment_type}


@router.get("/attachments/{filename}")
async def get_attachment(filename: str):
    file_path = UPLOADS_DIR / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    return StreamingResponse(open(file_path, "rb"), media_type="application/octet-stream", headers={"Content-Disposition": f"attachment; filename={filename}"})


# ================= EXPORT ENDPOINTS =================

@router.get("/export/transactions/excel")
async def export_transactions_excel(start_date: Optional[str] = None, end_date: Optional[str] = None, user: dict = Depends(get_current_user)):
    query = {"user_id": user["id"], "status": {"$ne": TransactionStatus.DUPLICATE_CONFIRMED}}
    if start_date:
        query["date"] = {"$gte": start_date}
    if end_date:
        query.setdefault("date", {})["$lte"] = end_date
    transactions = await db.transactions.find(query, {"_id": 0}).sort("date", -1).to_list(10000)
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)
    header_format = workbook.add_format({'bold': True, 'bg_color': '#2f9e44', 'font_color': 'white'})
    money_format = workbook.add_format({'num_format': '$#,##0.00'})
    ws1 = workbook.add_worksheet("Transacciones")
    headers = ["Fecha", "Descripcion", "Establecimiento", "Categoria", "Subcategoria", "Monto", "Tipo", "Deducible SRI", "Estado", "Fuente"]
    for col, header in enumerate(headers):
        ws1.write(0, col, header, header_format)
    for row, t in enumerate(transactions, 1):
        ws1.write(row, 0, t.get("date", ""))
        ws1.write(row, 1, t.get("description", ""))
        ws1.write(row, 2, t.get("establishment", ""))
        ws1.write(row, 3, SRI_CATEGORIES.get(t.get("category", ""), {}).get("name", t.get("category", "")))
        ws1.write(row, 4, t.get("subcategory", ""))
        ws1.write(row, 5, t.get("amount", 0), money_format)
        ws1.write(row, 6, "Ingreso" if t.get("transaction_type") == "income" else "Gasto")
        ws1.write(row, 7, "Si" if t.get("is_deductible") else "No")
        ws1.write(row, 8, STATUS_LABELS.get(t.get("status", ""), t.get("status", "")))
        ws1.write(row, 9, t.get("source_type", "manual"))
    ws1.autofilter(0, 0, len(transactions), len(headers) - 1)
    ws1.set_column(0, 0, 12)
    ws1.set_column(1, 1, 30)
    ws1.set_column(2, 4, 20)
    ws1.set_column(5, 5, 12)
    ws2 = workbook.add_worksheet("Resumen SRI")
    ws2.write(0, 0, "Categoria", header_format)
    ws2.write(0, 1, "Total", header_format)
    ws2.write(0, 2, "Deducible", header_format)
    ws2.write(0, 3, "Limite SRI", header_format)
    ws2.write(0, 4, "% Usado", header_format)
    summary = {}
    for t in transactions:
        if t.get("transaction_type") == "expense":
            cat = t.get("category", "otros")
            summary[cat] = summary.get(cat, 0) + t.get("amount", 0)
    row = 1
    for cat, total in summary.items():
        cat_info = SRI_CATEGORIES.get(cat, {})
        limit = cat_info.get("limit_usd", 0)
        ws2.write(row, 0, cat_info.get("name", cat))
        ws2.write(row, 1, total, money_format)
        ws2.write(row, 2, "Si" if cat_info.get("deductible") else "No")
        ws2.write(row, 3, limit, money_format)
        ws2.write(row, 4, f"{(total/limit*100):.1f}%" if limit > 0 else "N/A")
        row += 1
    workbook.close()
    output.seek(0)
    filename = f"transacciones_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/export/sri/pdf")
async def export_sri_pdf(year: Optional[int] = None, cargas_familiares: int = 3, user: dict = Depends(get_current_user)):
    if not year:
        year = datetime.now().year
    start_date = f"{year}-01-01"
    end_date = f"{year}-12-31"
    transactions = await db.transactions.find({"user_id": user["id"], "date": {"$gte": start_date, "$lte": end_date}, "transaction_type": "expense", "is_deductible": True, "uso_empresarial": {"$ne": True}, "status": {"$ne": TransactionStatus.DUPLICATE_CONFIRMED}}, {"_id": 0}).to_list(10000)
    category_totals = {}
    for t in transactions:
        cat = t.get("category", "otros")
        category_totals[cat] = category_totals.get(cat, 0) + t.get("amount", 0)
    cargas = min(cargas_familiares, 5)
    num_cbf = CARGAS_FAMILIARES_CBF.get(cargas, 7)
    limite_global = num_cbf * CANASTA_BASICA
    total_deductible = sum(category_totals.values())
    gastos_aplicables = min(total_deductible, limite_global)
    rebaja_ir = gastos_aplicables * PORCENTAJE_REBAJA_IR
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=16, spaceAfter=20)
    heading_style = ParagraphStyle('Heading', parent=styles['Heading2'], fontSize=12, spaceAfter=10)
    normal_style = styles['Normal']
    elements = []
    elements.append(Paragraph("ANEXO DE GASTOS PERSONALES", title_style))
    elements.append(Paragraph(f"Ano Fiscal: {year}", heading_style))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph("DATOS DEL CONTRIBUYENTE", heading_style))
    info_data = [["RUC:", CONTRIBUYENTE_INFO["ruc"]], ["Nombre:", CONTRIBUYENTE_INFO["nombre"]], ["Tipo:", CONTRIBUYENTE_INFO["tipo"]], ["Cargas Familiares:", str(cargas_familiares)]]
    info_table = Table(info_data, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'), ('ALIGN', (0, 0), (-1, -1), 'LEFT')]))
    elements.append(info_table)
    elements.append(Spacer(1, 20))
    elements.append(Paragraph("GASTOS DEDUCIBLES POR CATEGORIA", heading_style))
    table_data = [["Categoria", "Monto Gastado", "Limite", "% Usado"]]
    deductible_cats = ["alimentacion", "salud", "educacion", "vivienda", "vestimenta", "turismo"]
    for cat in deductible_cats:
        cat_info = SRI_CATEGORIES.get(cat, {})
        spent = category_totals.get(cat, 0)
        limit = cat_info.get("limit_usd", 0)
        pct = (spent / limit * 100) if limit > 0 else 0
        table_data.append([cat_info.get("name", cat), f"${spent:,.2f}", f"${limit:,.2f}", f"{pct:.1f}%"])
    table_data.append(["TOTAL DEDUCIBLE", f"${total_deductible:,.2f}", f"${limite_global:,.2f}", f"{(total_deductible/limite_global*100):.1f}%" if limite_global > 0 else "0%"])
    expenses_table = Table(table_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1*inch])
    expenses_table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2f9e44')), ('TEXTCOLOR', (0, 0), (-1, 0), colors.white), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'), ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#d3f9d8')), ('ALIGN', (1, 0), (-1, -1), 'RIGHT'), ('GRID', (0, 0), (-1, -1), 0.5, colors.grey), ('ROWHEIGHT', (0, 0), (-1, -1), 25)]))
    elements.append(expenses_table)
    elements.append(Spacer(1, 20))
    elements.append(Paragraph("CALCULO DE REBAJA", heading_style))
    rebaja_data = [["Gastos Aplicables (menor entre total y limite):", f"${gastos_aplicables:,.2f}"], ["Porcentaje de Rebaja:", f"{PORCENTAJE_REBAJA_IR*100:.0f}%"], ["REBAJA ESTIMADA DE IMPUESTO A LA RENTA:", f"${rebaja_ir:,.2f}"]]
    rebaja_table = Table(rebaja_data, colWidths=[4*inch, 2*inch])
    rebaja_table.setStyle(TableStyle([('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'), ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#2f9e44')), ('TEXTCOLOR', (0, -1), (-1, -1), colors.white), ('ALIGN', (1, 0), (1, -1), 'RIGHT'), ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)]))
    elements.append(rebaja_table)
    elements.append(Spacer(1, 30))
    elements.append(Paragraph(f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')} - FamilyFinance Ecuador", normal_style))
    elements.append(Paragraph("Este documento es un resumen para uso personal. Para la declaracion oficial, use el formulario del SRI.", normal_style))
    doc.build(elements)
    output.seek(0)
    filename = f"anexo_gastos_personales_{year}.pdf"
    return StreamingResponse(output, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})


# ================= ACCOUNTANT TAX SUMMARY =================

@router.get("/accountant/tax-summary")
async def get_tax_summary(year: Optional[int] = None, user: dict = Depends(check_role([UserRole.ADMIN, UserRole.ACCOUNTANT]))):
    if not year:
        year = datetime.now(timezone.utc).year
    start_date = f"{year}-01-01"
    end_date = f"{year}-12-31"
    transactions = await db.transactions.find({"date": {"$gte": start_date, "$lte": end_date}, "transaction_type": "expense"}, {"_id": 0}).to_list(10000)
    sri_deductible_cats = ["alimentacion", "salud", "educacion", "vivienda", "vestimenta"]
    summary = {"year": year, "total_expenses": 0, "deductible_expenses": 0, "by_category": {}, "monthly_breakdown": {}}
    for t in transactions:
        cat = t["category"]
        month = t["date"][:7]
        amount = t["amount"]
        summary["total_expenses"] += amount
        if cat in sri_deductible_cats:
            summary["deductible_expenses"] += amount
        if cat not in summary["by_category"]:
            summary["by_category"][cat] = {"name": SRI_CATEGORIES.get(cat, {}).get("name", cat), "total": 0, "deductible": cat in sri_deductible_cats}
        summary["by_category"][cat]["total"] += amount
        if month not in summary["monthly_breakdown"]:
            summary["monthly_breakdown"][month] = {"total": 0, "deductible": 0}
        summary["monthly_breakdown"][month]["total"] += amount
        if cat in sri_deductible_cats:
            summary["monthly_breakdown"][month]["deductible"] += amount
    return summary

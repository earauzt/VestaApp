from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Form, status
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
from passlib.context import CryptContext
from jose import JWTError, jwt
import json
import base64
import aiofiles
from emergentintegrations.llm.chat import LlmChat, UserMessage, FileContentWithMimeType
import tempfile
import openpyxl
from io import BytesIO
import re
import xlsxwriter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import pdfplumber

ROOT_DIR = Path(__file__).parent
UPLOADS_DIR = ROOT_DIR / "uploads"
UPLOADS_DIR.mkdir(exist_ok=True)

load_dotenv(ROOT_DIR / '.env')

# MongoDB connection with SSL certificate
import certifi
mongo_url = os.environ.get('MONGO_URL')
db_name = os.environ.get('DB_NAME')
client = AsyncIOMotorClient(mongo_url, tlsCAFile=certifi.where())
db = client[db_name]

# Import seed data function
from seed_data import seed_database

# JWT Settings
SECRET_KEY = os.environ.get('JWT_SECRET_KEY', 'default_secret_key')
ALGORITHM = os.environ.get('JWT_ALGORITHM', 'HS256')
ACCESS_TOKEN_EXPIRE_MINUTES = int(os.environ.get('ACCESS_TOKEN_EXPIRE_MINUTES', 1440))
EMERGENT_LLM_KEY = os.environ.get('EMERGENT_LLM_KEY', '')

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()

# Create the main app
app = FastAPI(title="FamilyFinance Ecuador API")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ================= MODELS =================

# User roles
class UserRole:
    ADMIN = "admin"
    ACCOUNTANT = "accountant"
    SPOUSE = "spouse"

# SRI Ecuador 2025 - Límites de deducción de Gastos Personales
# Según Ley de Régimen Tributario Interno Art. 10 numeral 16
# Resolución NAC-DGERCGC23-00000024

# Canasta Básica Familiar 2025 (referencia para límites SRI)
CANASTA_BASICA = 798.31  # USD Enero 2025

# Fracción Básica Exenta 2025
FRACCION_BASICA_EXENTA = 11902.00  # USD anual

# Tabla de límites por cargas familiares (número de CBF)
CARGAS_FAMILIARES_CBF = {
    0: 7,    # 0 cargas = 7 CBF = $5,588.17
    1: 9,    # 1 carga = 9 CBF = $7,184.79
    2: 11,   # 2 cargas = 11 CBF = $8,781.41
    3: 13,   # 3 cargas = 13 CBF = $10,378.03
    4: 15,   # 4 cargas = 15 CBF = $11,974.65
    5: 17,   # 5+ cargas = 17 CBF = $13,571.27
}

# Rebaja de Impuesto a la Renta: 18% de gastos deducibles
PORCENTAJE_REBAJA_IR = 0.18

# SRI Ecuador Categories - Límites de deducción 2025
# limit_percentage: Fracción de la Fracción Básica Exenta
SRI_CATEGORIES = {
    "alimentacion": {
        "name": "Alimentación", 
        "subcategories": ["Comida", "Restaurantes", "Supermercado", "Mercado"],
        "deductible": True, 
        "limit_percentage": 0.325,  # 0.325 * $11,902 = $3,868.15 máximo
        "limit_usd": 3868.15,
        "description": "Compras de alimentos, restaurantes, supermercados"
    },
    "salud": {
        "name": "Salud", 
        "subcategories": ["Seguros", "Medicina", "Consultas", "Hospitalización", "Laboratorio"],
        "deductible": True, 
        "limit_percentage": 1.3,  # 1.3 * $11,902 = $15,472.60 máximo (enfermedades catastróficas)
        "limit_usd": 15472.60,
        "description": "Consultas médicas, medicinas, seguros de salud, hospitalización"
    },
    "educacion": {
        "name": "Educación", 
        "subcategories": ["Colegio y actividades", "Cursos", "Materiales", "Universidad", "Maestría"],
        "deductible": True, 
        "limit_percentage": 0.325,
        "limit_usd": 3868.15,
        "description": "Matrículas, pensiones, útiles escolares, cursos, seminarios"
    },
    "vivienda": {
        "name": "Vivienda", 
        "subcategories": ["Servicios básicos", "Arriendo", "Intereses hipoteca", "Mantenimiento"],
        "deductible": True, 
        "limit_percentage": 0.325,
        "limit_usd": 3868.15,
        "description": "Arriendo, servicios básicos (agua, luz, teléfono), intereses hipotecarios"
    },
    "vestimenta": {
        "name": "Vestimenta", 
        "subcategories": ["Ropa", "Calzado", "Accesorios"],
        "deductible": True, 
        "limit_percentage": 0.325,
        "limit_usd": 3868.15,
        "description": "Ropa y calzado adquiridos en Ecuador"
    },
    "turismo": {
        "name": "Turismo Nacional",
        "subcategories": ["Hoteles Ecuador", "Tours locales", "Transporte turístico"],
        "deductible": True,
        "limit_percentage": 0.325,
        "limit_usd": 3868.15,
        "description": "Turismo dentro de Ecuador (hoteles, tours)"
    },
    "transporte": {
        "name": "Transporte", 
        "subcategories": ["Carros", "Combustible", "Mantenimiento vehicular", "Taxi", "Bus"],
        "deductible": False, 
        "limit_percentage": 0,
        "limit_usd": 0,
        "description": "NO DEDUCIBLE - Transporte personal"
    },
    "viajes_internacionales": {
        "name": "Viajes Internacionales", 
        "subcategories": ["USA", "Europa", "Otros países"],
        "deductible": False, 
        "limit_percentage": 0,
        "limit_usd": 0,
        "description": "NO DEDUCIBLE - Gastos en el exterior"
    },
    "otros": {
        "name": "Otros", 
        "subcategories": ["Empleados", "Entretenimiento", "Varios"],
        "deductible": False, 
        "limit_percentage": 0,
        "limit_usd": 0,
        "description": "NO DEDUCIBLE - Gastos varios"
    }
}

# Payment methods (auto-detected)
PAYMENT_METHODS = {
    "transferencia": {"name": "Transferencia", "keywords": ["transfer", "wire", "venmo", "zelle", "deposito", "banco"]},
    "tarjeta": {"name": "Tarjeta", "keywords": ["visa", "mastercard", "card", "tarjeta", "pacificard", "diners"]},
    "efectivo": {"name": "Efectivo", "keywords": ["cash", "efectivo", "contado"]},
    "venmo": {"name": "Venmo", "keywords": ["venmo"]},
    "apple_card": {"name": "Apple Card", "keywords": ["apple card", "apple pay", "apple cash"]}
}

# Payment sources
PAYMENT_SOURCES = ["local", "internacional"]  # tarjeta local vs tarjeta extranjera

# Income Sources - Distribution for tax purposes
INCOME_SOURCES = ["Personal", "APX", "USA"]

# Income Concepts
INCOME_CONCEPTS = ["Salario", "Bonus", "Dividendos", "Arriendo", "Honorarios", "Otros"]

# ================= PERSONAL BUDGET CATEGORIES (User's Excel - PROYECCION EA 2026) =================
# Extracted from user's Excel screenshots - exact monthly budgets
BUDGET_CATEGORIES = {
    "servicios_basicos": {
        "name": "Servicios Básicos",
        "subcategories": {
            "Alícuota B": 500,
            "Alícuota GT": 100,
            "Luz": 200,
            "Gas": 15,
            "Celular": 80,
            "Agua": 60,
            "Clubes": 275,
            "Internet": 50,
            "Suscripciones": 0
        },
        "monthly_budget": 1280,
        "annual_budget": 15360,
        "type": "fixed",
        "payment_methods": ["transferencia", "tarjeta"],
        "is_recurring": True
    },
    "suscripciones": {
        "name": "Suscripciones",
        "subcategories": {
            "Netflix": 0,
            "Spotify": 0,
            "Amazon Prime": 0,
            "Disney+": 0,
            "YouTube Premium": 0,
            "iCloud": 0,
            "Otras": 0
        },
        "monthly_budget": 0,
        "annual_budget": 0,
        "type": "recurring",
        "payment_methods": ["tarjeta"],
        "is_recurring": True,
        "tags": ["recurrente", "suscripcion"]
    },
    "empleados": {
        "name": "Empleados",
        "subcategories": {
            "Ramona": 600,
            "Angélica": 550,
            "IESS": 150
        },
        "monthly_budget": 1300,
        "annual_budget": 15600,
        "type": "fixed",
        "payment_methods": ["transferencia", "efectivo"],
        "is_recurring": True
    },
    "colegio_actividades": {
        "name": "Colegio y Actividades",
        "subcategories": {
            "Menor": 2000,
            "Fútbol": 150,
            "Telas Aros": 210
        },
        "monthly_budget": 2360,
        "annual_budget": 28320,
        "type": "fixed",
        "payment_methods": ["transferencia", "tarjeta", "efectivo"],
        "is_recurring": True
    },
    "seguros": {
        "name": "Seguros",
        "subcategories": {
            "Salud": 900,
            "Carros": 250
        },
        "monthly_budget": 1150,
        "annual_budget": 13800,
        "type": "fixed",
        "payment_methods": ["tarjeta", "transferencia"],
        "is_recurring": True
    },
    "comida": {
        "name": "Comida",
        "subcategories": {
            "Supermaxi": 800,
            "Mercado": 150
        },
        "monthly_budget": 950,
        "annual_budget": 11400,
        "type": "variable",
        "payment_methods": ["tarjeta", "efectivo"],
        "is_recurring": False
    },
    "restaurantes": {
        "name": "Restaurantes",
        "subcategories": {
            "Comida afuera": 350,
            "Delivery": 200
        },
        "monthly_budget": 550,
        "annual_budget": 6600,
        "type": "variable",
        "payment_methods": ["tarjeta", "efectivo"],
        "is_recurring": False
    },
    "carros": {
        "name": "Carros",
        "subcategories": {
            "Gasolina 1": 360,
            "Gasolina 2": 105,
            "Mantenimiento": 100
        },
        "monthly_budget": 565,
        "annual_budget": 6780,
        "type": "variable",
        "payment_methods": ["tarjeta", "efectivo"],
        "is_recurring": False
    },
    "usa": {
        "name": "USA",
        "subcategories": {
            "Mamá (Venmo)": 600,
            "TMobile": 150,
            "Universidad": 400
        },
        "monthly_budget": 1150,
        "annual_budget": 13800,
        "type": "fixed",
        "payment_methods": ["venmo", "apple_card", "transferencia"],
        "is_recurring": True,
        "is_international": True
    },
    "viajes_entretenimiento": {
        "name": "Viajes y Entretenimiento",
        "subcategories": {
            "Pasajes": 500,
            "Navidad": 7000
        },
        "monthly_budget": 0,
        "annual_budget": 16500,
        "type": "variable",
        "payment_methods": ["tarjeta", "apple_card"],
        "is_recurring": False,
        "notes": "Pasajes en Enero $500, Diciembre $3000. Navidad $7000 en Diciembre"
    },
    "gastos_libres": {
        "name": "Gastos Libres (Otros)",
        "subcategories": {
            "KP (Esposa)": 800,
            "EA (Emilio)": 500
        },
        "monthly_budget": 1300,
        "annual_budget": 15600,
        "type": "discretionary",
        "payment_methods": ["tarjeta", "efectivo"],
        "is_recurring": True
    }
}

# Known subscription services to auto-detect
SUBSCRIPTION_SERVICES = [
    "netflix", "spotify", "amazon prime", "disney", "hbo", "youtube", 
    "apple music", "icloud", "google one", "dropbox", "adobe", 
    "microsoft 365", "office 365", "chatgpt", "openai", "canva",
    "audible", "kindle", "paramount", "star+", "crunchyroll"
]

# Income structure from Excel (FLUJO AÑO 2026) - ESTIMATED/VARIABLE
INCOME_STRUCTURE = {
    "personal": {"monthly": 7250, "annual": 87000, "source": "Personal", "note": "Estimado variable"},
    "apx": {"monthly": 2500, "annual": 30000, "source": "APX", "note": "Estimado variable"},
    "usa": {"monthly": 2750, "annual": 33000, "source": "USA", "note": "Estimado variable"}
}
TOTAL_MONTHLY_INCOME = 12500  # Estimado
TOTAL_ANNUAL_INCOME = 150000  # Estimado

# Budget summary from Excel
BUDGET_SUMMARY = {
    "total_gastos_fijos_monthly": 8155,  # Sum of fixed categories
    "total_gastos_fijos_annual": 97860,
    "flujo_neto_mensual": 1595,  # After Ecuador expenses
    "ahorro_esperado": {"monthly": 1250, "annual": 15000, "percentage": 10},
    "inversion_esperada": {"monthly": 1875, "annual": 22500, "percentage": 15}
}

# Budget Goals (User's financial targets)
BUDGET_GOALS = {
    "gastos_fijos_target": {"min": 0.55, "max": 0.65, "name": "Gastos Fijos"},
    "ahorro_target": {"min": 0.10, "max": 0.10, "name": "Ahorro"},
    "inversion_target": {"min": 0.15, "max": 0.15, "name": "Inversión"},
    "gastos_libres_max_annual": 30000  # Max $30k/year
}

# Countries considered international
INTERNATIONAL_COUNTRIES = ["USA", "United States", "Estados Unidos", "US", "EU", "Europa", "Spain", "España", "Colombia", "Peru", "Perú", "México", "Mexico", "Miami", "New York", "Los Angeles", "Houston", "Texas", "California", "Florida"]

# ================= DEMO USER DATA =================
# Datos ficticios para usuario demo (no muestra datos reales)
DEMO_BUDGET_CATEGORIES = {
    "servicios_basicos": {
        "name": "Servicios Básicos",
        "subcategories": {"Luz": 60, "Agua": 25, "Internet": 45, "Gas": 15},
        "monthly_budget": 145,
        "annual_budget": 1740,
        "type": "fixed"
    },
    "comida": {
        "name": "Alimentación",
        "subcategories": {"Supermercado": 300, "Mercado": 50},
        "monthly_budget": 350,
        "annual_budget": 4200,
        "type": "variable"
    },
    "restaurantes": {
        "name": "Restaurantes",
        "subcategories": {"Restaurantes": 150, "Delivery": 50},
        "monthly_budget": 200,
        "annual_budget": 2400,
        "type": "discretionary"
    },
    "transporte": {
        "name": "Transporte",
        "subcategories": {"Gasolina": 80, "Mantenimiento": 40},
        "monthly_budget": 120,
        "annual_budget": 1440,
        "type": "variable"
    },
    "entretenimiento": {
        "name": "Entretenimiento",
        "subcategories": {"Streaming": 30, "Salidas": 70},
        "monthly_budget": 100,
        "annual_budget": 1200,
        "type": "discretionary"
    },
    "otros": {
        "name": "Otros Gastos",
        "subcategories": {"Varios": 100},
        "monthly_budget": 100,
        "annual_budget": 1200,
        "type": "discretionary"
    }
}

DEMO_INCOME_STRUCTURE = {
    "personal": {"monthly": 3500, "annual": 42000, "source": "Salario", "note": "Ingreso principal"}
}

DEMO_BUDGET_SUMMARY = {
    "total_gastos_fijos_monthly": 1015,
    "total_gastos_fijos_annual": 12180,
    "flujo_neto_mensual": 2485,
    "ahorro_esperado": {"monthly": 350, "annual": 4200, "percentage": 10},
    "inversion_esperada": {"monthly": 175, "annual": 2100, "percentage": 5}
}

DEMO_BUDGET_GOALS = {
    "gastos_fijos_target": {"min": 0.25, "max": 0.35, "name": "Gastos Fijos"},
    "ahorro_target": {"min": 0.10, "max": 0.10, "name": "Ahorro"},
    "inversion_target": {"min": 0.05, "max": 0.05, "name": "Inversión"},
    "gastos_libres_max_annual": 5000
}

DEMO_CONTRIBUYENTE_INFO = {
    "ruc": "0900000000001",
    "nombre": "USUARIO DEMO",
    "tipo": "PERSONA NATURAL",
    "regimen": "RIMPE",
    "obligado_contabilidad": False,
    "actividad_principal": "SERVICIOS PROFESIONALES",
    "jurisdiccion": "GUAYAS / GUAYAQUIL",
    "cargas_familiares": 1,
    "cargas_detalle": [{"tipo": "conyuge", "nombre": "Cónyuge Demo"}]
}

def is_demo_user(user: dict) -> bool:
    """Check if user is a demo user"""
    return user.get("role") == "demo" or user.get("email") == "demo@fintrack.ec"

def get_budget_categories(user: dict):
    """Get budget categories based on user type"""
    return DEMO_BUDGET_CATEGORIES if is_demo_user(user) else BUDGET_CATEGORIES

def get_income_structure(user: dict):
    """Get income structure based on user type"""
    return DEMO_INCOME_STRUCTURE if is_demo_user(user) else INCOME_STRUCTURE

def get_budget_summary(user: dict):
    """Get budget summary based on user type"""
    return DEMO_BUDGET_SUMMARY if is_demo_user(user) else BUDGET_SUMMARY

def get_budget_goals(user: dict):
    """Get budget goals based on user type"""
    return DEMO_BUDGET_GOALS if is_demo_user(user) else BUDGET_GOALS

def get_contribuyente_info(user: dict):
    """Get contribuyente info based on user type"""
    return DEMO_CONTRIBUYENTE_INFO if is_demo_user(user) else CONTRIBUYENTE_INFO

# Datos del contribuyente (extraídos del RUC)
CONTRIBUYENTE_INFO = {
    "ruc": "0912514890001",
    "nombre": "ARAUZ TRIVIÑO EMILIO JOSE",
    "tipo": "PERSONA NATURAL",
    "regimen": "GENERAL",
    "obligado_contabilidad": False,
    "actividad_principal": "SERVICIOS DE MARKETING Y PUBLICIDAD",
    "jurisdiccion": "ZONA 8 / GUAYAS / SAMBORONDON",
    "cargas_familiares": 3,  # 2 hijos menores + esposa
    "cargas_detalle": [
        {"tipo": "conyuge", "nombre": "Esposa"},
        {"tipo": "hijo", "nombre": "Hijo 1"},
        {"tipo": "hijo", "nombre": "Hijo 2"}
    ]
}

# ================= AUTO-CATEGORIZATION RULES =================
# Reglas automáticas de categorización (estilo QuickBooks)
DEFAULT_CATEGORIZATION_RULES = [
    # Alimentación - Supermercados y restaurantes Ecuador
    {"keywords": ["supermaxi", "mi comisariato", "megamaxi", "tia", "aki", "gran aki", "coral"], "category": "alimentacion", "subcategory": "Supermercado"},
    {"keywords": ["mcdonalds", "mcdonald's", "burger king", "kfc", "pollo", "pizza hut", "dominos", "subway", "juan valdez", "sweet & coffee"], "category": "alimentacion", "subcategory": "Restaurantes"},
    {"keywords": ["mercado", "feria", "verduras", "frutas", "carniceria", "panaderia"], "category": "alimentacion", "subcategory": "Comida"},
    
    # Salud
    {"keywords": ["farmacia", "fybeca", "pharmacy", "medicity", "cruz azul", "sana sana", "economicas"], "category": "salud", "subcategory": "Medicina"},
    {"keywords": ["hospital", "clinica", "consultorio", "medico", "doctor", "laboratorio", "examen"], "category": "salud", "subcategory": "Consultas"},
    {"keywords": ["seguro medico", "seguros", "salud sa", "bmi", "humana", "saludsa", "ecuasanitas"], "category": "salud", "subcategory": "Seguros"},
    
    # Educación
    {"keywords": ["colegio", "escuela", "liceo", "unidad educativa", "academia"], "category": "educacion", "subcategory": "Colegio y actividades"},
    {"keywords": ["universidad", "uees", "espol", "ucsg", "usfq", "udla", "maestria", "postgrado"], "category": "educacion", "subcategory": "Universidad"},
    {"keywords": ["curso", "capacitacion", "udemy", "coursera", "platzi", "taller"], "category": "educacion", "subcategory": "Cursos"},
    {"keywords": ["libreria", "libro", "papeleria", "utiles"], "category": "educacion", "subcategory": "Materiales"},
    
    # Vivienda
    {"keywords": ["luz", "electrica", "cnel", "energia"], "category": "vivienda", "subcategory": "Servicios básicos"},
    {"keywords": ["agua potable", "interagua", "emapag"], "category": "vivienda", "subcategory": "Servicios básicos"},
    {"keywords": ["telefono", "cnt", "claro", "movistar", "internet", "netlife", "tv cable"], "category": "vivienda", "subcategory": "Servicios básicos"},
    {"keywords": ["arriendo", "alquiler", "renta mensual"], "category": "vivienda", "subcategory": "Arriendo"},
    {"keywords": ["hipoteca", "credito hipotecario", "banco vivienda"], "category": "vivienda", "subcategory": "Intereses hipoteca"},
    
    # Vestimenta
    {"keywords": ["zara", "h&m", "forever 21", "mango", "tennis", "etafashion", "de prati", "ri", "payless"], "category": "vestimenta", "subcategory": "Ropa"},
    {"keywords": ["marathon", "nike", "adidas", "puma", "calzado", "zapatos"], "category": "vestimenta", "subcategory": "Calzado"},
    
    # Transporte (NO deducible)
    {"keywords": ["gasolina", "diesel", "primax", "mobil", "petroecuador", "terpel", "combustible"], "category": "transporte", "subcategory": "Combustible"},
    {"keywords": ["mecanica", "taller", "llantas", "aceite motor", "repuestos"], "category": "transporte", "subcategory": "Mantenimiento vehicular"},
    {"keywords": ["uber", "cabify", "taxi", "indriver"], "category": "transporte", "subcategory": "Taxi"},
    
    # Turismo Nacional
    {"keywords": ["hotel ecuador", "hostal", "airbnb ecuador", "decameron", "hilton colon"], "category": "turismo", "subcategory": "Hoteles Ecuador"},
    
    # Viajes Internacionales (NO deducible)
    {"keywords": ["amazon.com", "ebay", "aliexpress", "wish", "shein"], "category": "viajes_internacionales", "subcategory": "USA"},
    {"keywords": ["booking.com internacional", "expedia", "hotel usa", "hotel miami"], "category": "viajes_internacionales", "subcategory": "USA"},
    
    # Otros (NO deducible)
    {"keywords": ["netflix", "spotify", "disney", "hbo", "prime video", "youtube premium"], "category": "otros", "subcategory": "Entretenimiento"},
    {"keywords": ["empleada", "domestico", "jardinero", "limpieza casa"], "category": "otros", "subcategory": "Empleados"},
]

def apply_categorization_rules(description: str, establishment: str = "") -> dict:
    """Apply automatic categorization rules based on keywords"""
    text = f"{description} {establishment}".lower()
    
    for rule in DEFAULT_CATEGORIZATION_RULES:
        for keyword in rule["keywords"]:
            if keyword.lower() in text:
                return {
                    "category": rule["category"],
                    "subcategory": rule["subcategory"],
                    "auto_categorized": True,
                    "matched_keyword": keyword
                }
    
    return {"category": None, "subcategory": None, "auto_categorized": False, "matched_keyword": None}

def detect_payment_method(description: str, establishment: str = "") -> str:
    """Auto-detect payment method from description"""
    text = f"{description} {establishment}".lower()
    
    for method_key, method_info in PAYMENT_METHODS.items():
        for keyword in method_info["keywords"]:
            if keyword.lower() in text:
                return method_key
    
    return "tarjeta"  # Default to card

# Transaction Status (inspirado en QuickBooks)
class TransactionStatus:
    PENDING_REVIEW = "pending_review"  # Pendiente de revisión por contadora
    APPROVED = "approved"  # Aprobado/conciliado
    REJECTED = "rejected"  # Rechazado (error o inválido)
    DUPLICATE_SUSPECT = "duplicate_suspect"  # Sospecha de duplicado
    DUPLICATE_CONFIRMED = "duplicate_confirmed"  # Confirmado como duplicado (no contar)

# Source Type (de dónde viene la transacción)
class SourceType:
    MANUAL = "manual"  # Ingresado manualmente
    EMAIL = "email"  # Desde email de tarjeta
    RECEIPT = "receipt"  # Desde foto de recibo
    INVOICE = "invoice"  # Desde factura
    EXCEL = "excel"  # Desde Excel
    BANK_STATEMENT = "bank_statement"  # Desde estado de cuenta

# Pydantic Models
class UserBase(BaseModel):
    email: EmailStr
    name: str
    role: str = UserRole.SPOUSE

class UserCreate(UserBase):
    password: str

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    email: str
    name: str
    role: str
    created_at: str

class Token(BaseModel):
    access_token: str
    token_type: str
    user: UserResponse

class TransactionBase(BaseModel):
    amount: float
    description: str
    category: str
    subcategory: Optional[str] = None
    date: str
    transaction_type: str = "expense"  # expense or income
    source: Optional[str] = None  # For income: Personal, APX, USA
    establishment: Optional[str] = None
    card_last_digits: Optional[str] = None
    country: Optional[str] = None  # País del gasto
    is_international: bool = False  # Si es gasto internacional
    payment_source: str = "local"  # local o internacional (tarjeta extranjera)
    is_deductible: bool = True  # Si es deducible para SRI
    # Nuevos campos para conciliación (estilo QuickBooks)
    status: str = TransactionStatus.PENDING_REVIEW  # Estado de revisión
    source_type: str = SourceType.MANUAL  # De dónde viene la transacción
    has_receipt: bool = False  # Tiene recibo/foto
    has_invoice: bool = False  # Tiene factura electrónica
    invoice_number: Optional[str] = None  # Número de factura
    notes: Optional[str] = None  # Notas adicionales
    reviewed_by: Optional[str] = None  # ID del usuario que revisó
    reviewed_at: Optional[str] = None  # Fecha de revisión
    duplicate_of: Optional[str] = None  # ID de la transacción original si es duplicado
    match_confidence: Optional[float] = None  # Confianza de match (0-100)
    # Campos para split y adjuntos
    is_split: bool = False  # Si es parte de un split
    parent_transaction_id: Optional[str] = None  # ID de la transacción padre (para splits)
    attachments: List[str] = []  # Lista de URLs/paths de archivos adjuntos
    auto_categorized: bool = False  # Si fue categorizado automáticamente
    matched_rule: Optional[str] = None  # Keyword que hizo match en auto-categorización
    # NEW: Payment method and budget category
    payment_method: Optional[str] = None  # transferencia, tarjeta, efectivo, venmo, apple_card
    budget_category: Optional[str] = None  # Personal budget category (different from SRI)
    receipt_group_id: Optional[str] = None  # Group transactions from same receipt/invoice

class TransactionCreate(TransactionBase):
    pass

class TransactionResponse(TransactionBase):
    model_config = ConfigDict(extra="ignore")
    id: str
    user_id: str
    created_at: str
    ai_classified: bool = False

# Modelo para split de transacciones
class SplitItem(BaseModel):
    amount: float
    category: str
    subcategory: str
    description: Optional[str] = None

class TransactionSplitRequest(BaseModel):
    transaction_id: str
    splits: List[SplitItem]

# Modelo para reglas de categorización
class CategorizationRule(BaseModel):
    keywords: List[str]
    category: str
    subcategory: str
    is_active: bool = True

# NEW: Income Entry Model (Manual entry with distribution)
class IncomeEntry(BaseModel):
    amount: float
    date: str
    distribution: str  # Personal, APX, USA
    concept: str  # Salario, Bonus, Dividendos, etc.
    description: Optional[str] = None
    is_recurring: bool = False
    payment_method: Optional[str] = None

class IncomeResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    user_id: str
    amount: float
    date: str
    distribution: str
    concept: str
    description: Optional[str] = None
    is_recurring: bool
    payment_method: Optional[str] = None
    created_at: str

class BudgetItem(BaseModel):
    category: str
    subcategory: str
    planned_amount: float
    month: str  # Format: YYYY-MM

class BudgetResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    user_id: str
    items: List[BudgetItem]
    total_income: float
    total_expenses: float
    created_at: str

# ================= CREDIT CARDS & DEBT MODELS =================

class CreditCard(BaseModel):
    name: str  # Diners, Pichincha, Pacificard, Apple Card
    bank: str
    apr: float  # Annual Percentage Rate (15% Ecuador, 29% USA)
    credit_limit: float
    current_balance: float = 0
    minimum_payment: float = 0
    cut_off_day: int  # Day of month (1-31)
    payment_due_day: int  # Day of month for payment
    currency: str = "USD"
    is_international: bool = False

class CreditCardResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    user_id: str
    name: str
    bank: str
    apr: float
    credit_limit: float
    current_balance: float
    minimum_payment: float
    available_credit: float
    cut_off_day: int
    payment_due_day: int
    currency: str
    is_international: bool
    created_at: str
    updated_at: Optional[str] = None

class DebtPayment(BaseModel):
    card_id: str
    amount: float
    date: str
    payment_type: str  # minimum, full, custom

class SnowballPlan(BaseModel):
    strategy: str  # avalanche (highest interest first) or snowball (smallest balance first)
    extra_payment: float  # Extra amount to pay beyond minimums
    
# ================= CASH FLOW PLANNING MODELS =================

class ScheduledPayment(BaseModel):
    name: str  # "Luz", "Internet", "Colegio"
    amount: float
    due_day: int  # Day of month
    category: str
    payment_method: str  # transferencia, tarjeta_diners, tarjeta_pichincha, etc.
    is_recurring: bool = True
    reminder_days_before: int = 2  # Days before to remind

class ScheduledPaymentResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    user_id: str
    name: str
    amount: float
    due_day: int
    category: str
    payment_method: str
    is_recurring: bool
    reminder_days_before: int
    last_paid_date: Optional[str] = None
    next_due_date: str
    created_at: str

# ================= EXPECTED INCOME & RECEIVABLES MODELS =================

class ExpectedIncomeCreate(BaseModel):
    description: str
    amount: float
    expected_date: str
    source: str = "personal"  # personal, apx, usa
    recurring: bool = False
    recurring_frequency: Optional[str] = None  # monthly, biweekly
    notes: Optional[str] = None

class ExpectedIncomeResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    user_id: str
    description: str
    amount: float
    expected_date: str
    source: str
    status: str  # pending, received, cancelled
    recurring: bool
    recurring_frequency: Optional[str] = None
    linked_transaction_id: Optional[str] = None
    notes: Optional[str] = None
    created_at: str

class AccountReceivableCreate(BaseModel):
    client_name: str
    invoice_number: Optional[str] = None
    amount: float
    invoice_date: str
    due_date: str
    notes: Optional[str] = None

class AccountReceivableResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    user_id: str
    client_name: str
    invoice_number: Optional[str] = None
    amount: float
    amount_paid: float
    invoice_date: str
    due_date: str
    status: str  # pending, overdue, paid, partial
    notes: Optional[str] = None
    payment_history: List[dict] = []
    created_at: str

class TravelGoalCreate(BaseModel):
    destination: str
    target_amount: float
    target_date: str
    notes: Optional[str] = None

class TravelGoalResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    user_id: str
    destination: str
    target_amount: float
    saved_amount: float
    target_date: str
    status: str  # active, completed, cancelled
    notes: Optional[str] = None
    created_at: str

class DocumentUpload(BaseModel):
    document_type: str  # email, receipt, statement, excel
    content: Optional[str] = None

class PredictionResponse(BaseModel):
    category: str
    current_spending: float
    predicted_spending: float
    trend: str
    advice: str

class DashboardStats(BaseModel):
    total_income: float
    total_expenses: float
    balance: float
    daily_average: float
    weekly_total: float
    monthly_total: float
    by_category: Dict[str, float]
    sri_deductible: float

# ================= AUTH HELPERS =================

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password: str) -> str:
    return pwd_context.hash(password)

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + (expires_delta or timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Token inválido")
    except JWTError:
        raise HTTPException(status_code=401, detail="Token inválido")
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if user is None:
        raise HTTPException(status_code=401, detail="Usuario no encontrado")
    return user

def check_role(allowed_roles: List[str]):
    async def role_checker(user: dict = Depends(get_current_user)):
        if user["role"] not in allowed_roles:
            raise HTTPException(status_code=403, detail="Acceso denegado")
        return user
    return role_checker

# ================= AUTH ENDPOINTS =================

@api_router.post("/auth/register", response_model=Token)
async def register(user_data: UserCreate):
    # Check if user exists
    existing = await db.users.find_one({"email": user_data.email})
    if existing:
        raise HTTPException(status_code=400, detail="El email ya está registrado")
    
    # Create user
    user_id = str(uuid.uuid4())
    hashed_password = get_password_hash(user_data.password)
    
    user_doc = {
        "id": user_id,
        "email": user_data.email,
        "name": user_data.name,
        "role": user_data.role,
        "password": hashed_password,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.users.insert_one(user_doc)
    
    # Create token
    access_token = create_access_token(data={"sub": user_id})
    
    return Token(
        access_token=access_token,
        token_type="bearer",
        user=UserResponse(
            id=user_id,
            email=user_data.email,
            name=user_data.name,
            role=user_data.role,
            created_at=user_doc["created_at"]
        )
    )

@api_router.post("/auth/login", response_model=Token)
async def login(credentials: UserLogin):
    user = await db.users.find_one({"email": credentials.email}, {"_id": 0})
    if not user or not verify_password(credentials.password, user.get("hashed_password", user.get("password", ""))):
        raise HTTPException(status_code=401, detail="Credenciales inválidas")
    
    access_token = create_access_token(data={"sub": user["id"]})
    
    return Token(
        access_token=access_token,
        token_type="bearer",
        user=UserResponse(
            id=user["id"],
            email=user["email"],
            name=user["name"],
            role=user["role"],
            created_at=user["created_at"]
        )
    )

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(user: dict = Depends(get_current_user)):
    return UserResponse(
        id=user["id"],
        email=user["email"],
        name=user["name"],
        role=user["role"],
        created_at=user["created_at"]
    )

# ================= CATEGORIES ENDPOINTS =================

@api_router.get("/categories")
async def get_categories():
    return {
        "categories": SRI_CATEGORIES, 
        "income_sources": INCOME_SOURCES,
        "payment_sources": PAYMENT_SOURCES,
        "international_countries": INTERNATIONAL_COUNTRIES,
        "canasta_basica": CANASTA_BASICA,
        "fraccion_basica_exenta": FRACCION_BASICA_EXENTA,
        "contribuyente": CONTRIBUYENTE_INFO
    }

@api_router.get("/sri/deduction-limits")
async def get_sri_deduction_limits(
    cargas_familiares: int = 0,
    user: dict = Depends(get_current_user)
):
    """Get SRI deduction limits based on family dependents"""
    # Get current year transactions
    now = datetime.now(timezone.utc)
    start_of_year = f"{now.year}-01-01"
    
    transactions = await db.transactions.find(
        {"user_id": user["id"], "date": {"$gte": start_of_year}, "transaction_type": "expense"},
        {"_id": 0}
    ).to_list(10000)
    
    # Calculate spent by category
    spent_by_category = {}
    total_deductible = 0
    total_non_deductible = 0
    
    for t in transactions:
        cat = t.get("category", "otros")
        amount = t.get("amount", 0)
        
        if cat not in spent_by_category:
            spent_by_category[cat] = 0
        spent_by_category[cat] += amount
        
        if SRI_CATEGORIES.get(cat, {}).get("deductible", False):
            total_deductible += amount
        else:
            total_non_deductible += amount
    
    # Calculate limits
    cargas = min(cargas_familiares, 5)
    num_cbf = CARGAS_FAMILIARES_CBF.get(cargas, 7)
    limite_global = num_cbf * CANASTA_BASICA
    
    # Calculate progress for each category
    category_progress = []
    for cat_key, cat_info in SRI_CATEGORIES.items():
        if cat_info.get("deductible", False):
            spent = spent_by_category.get(cat_key, 0)
            limit = cat_info.get("limit_usd", 0)
            percentage = (spent / limit * 100) if limit > 0 else 0
            
            category_progress.append({
                "category": cat_key,
                "name": cat_info["name"],
                "spent": round(spent, 2),
                "limit": limit,
                "percentage": round(min(percentage, 100), 1),
                "remaining": round(max(0, limit - spent), 2),
                "over_limit": spent > limit,
                "description": cat_info.get("description", "")
            })
    
    # Sort by percentage (highest first)
    category_progress.sort(key=lambda x: x["percentage"], reverse=True)
    
    # Calculate potential tax rebate
    gastos_aplicables = min(total_deductible, limite_global)
    rebaja_ir = gastos_aplicables * PORCENTAJE_REBAJA_IR
    
    return {
        "year": now.year,
        "contribuyente": CONTRIBUYENTE_INFO,
        "cargas_familiares": cargas_familiares,
        "canasta_basica": CANASTA_BASICA,
        "fraccion_basica_exenta": FRACCION_BASICA_EXENTA,
        "limite_global": round(limite_global, 2),
        "total_deductible_spent": round(total_deductible, 2),
        "total_non_deductible_spent": round(total_non_deductible, 2),
        "gastos_aplicables": round(gastos_aplicables, 2),
        "rebaja_ir_estimada": round(rebaja_ir, 2),
        "porcentaje_rebaja": PORCENTAJE_REBAJA_IR * 100,
        "percentage_used": round((total_deductible / limite_global * 100) if limite_global > 0 else 0, 1),
        "remaining_global": round(max(0, limite_global - total_deductible), 2),
        "category_progress": category_progress,
        "alerts": _generate_sri_alerts(category_progress, total_deductible, limite_global)
    }

def _generate_sri_alerts(category_progress, total_deductible, limite_global):
    """Generate alerts for SRI deductions"""
    alerts = []
    
    # Check global limit
    if total_deductible >= limite_global * 0.9:
        alerts.append({
            "type": "warning",
            "message": f"Has usado el 90% de tu límite global de deducciones (${limite_global:,.2f})"
        })
    
    # Check individual categories
    for cat in category_progress:
        if cat["percentage"] >= 100:
            alerts.append({
                "type": "error",
                "message": f"LÍMITE EXCEDIDO en {cat['name']}: ${cat['spent']:,.2f} de ${cat['limit']:,.2f}"
            })
        elif cat["percentage"] >= 80:
            alerts.append({
                "type": "warning", 
                "message": f"{cat['name']}: {cat['percentage']}% del límite usado. Quedan ${cat['remaining']:,.2f}"
            })
    
    return alerts

# ================= DUPLICATE DETECTION =================

async def find_potential_duplicates(user_id: str, amount: float, date: str, establishment: str = None, description: str = None):
    """Find potential duplicate transactions (estilo QuickBooks)"""
    # Search for transactions with same amount within 7 days
    date_obj = datetime.strptime(date, "%Y-%m-%d")
    date_start = (date_obj - timedelta(days=7)).strftime("%Y-%m-%d")
    date_end = (date_obj + timedelta(days=7)).strftime("%Y-%m-%d")
    
    query = {
        "user_id": user_id,
        "amount": {"$gte": amount * 0.95, "$lte": amount * 1.05},  # 5% tolerance
        "date": {"$gte": date_start, "$lte": date_end},
        "status": {"$nin": [TransactionStatus.DUPLICATE_CONFIRMED, TransactionStatus.REJECTED]}
    }
    
    potential_duplicates = await db.transactions.find(query, {"_id": 0}).to_list(10)
    
    duplicates = []
    for t in potential_duplicates:
        confidence = 50  # Base confidence for same amount + date range
        
        # Increase confidence based on matching fields
        if establishment and t.get("establishment"):
            if establishment.lower() in t["establishment"].lower() or t["establishment"].lower() in establishment.lower():
                confidence += 30
        
        if description and t.get("description"):
            # Simple word matching
            words1 = set(description.lower().split())
            words2 = set(t["description"].lower().split())
            common_words = words1.intersection(words2)
            if len(common_words) >= 2:
                confidence += 20
        
        # Exact amount match
        if t.get("amount") == amount:
            confidence += 10
        
        if confidence >= 60:  # Threshold for flagging
            duplicates.append({
                "transaction": t,
                "confidence": min(confidence, 100)
            })
    
    return duplicates

# ================= TRANSACTIONS ENDPOINTS =================

@api_router.post("/transactions", response_model=TransactionResponse)
async def create_transaction(
    transaction: TransactionCreate,
    user: dict = Depends(get_current_user)
):
    transaction_id = str(uuid.uuid4())
    
    # Check for potential duplicates
    duplicates = await find_potential_duplicates(
        user["id"], 
        transaction.amount, 
        transaction.date,
        transaction.establishment,
        transaction.description
    )
    
    # Determine initial status
    initial_status = transaction.status
    duplicate_of = None
    match_confidence = None
    
    if duplicates:
        # Mark as duplicate suspect
        initial_status = TransactionStatus.DUPLICATE_SUSPECT
        duplicate_of = duplicates[0]["transaction"]["id"]
        match_confidence = duplicates[0]["confidence"]
    
    # Check if international based on country
    is_international = transaction.is_international
    if transaction.country:
        is_international = any(c.lower() in transaction.country.lower() for c in INTERNATIONAL_COUNTRIES)
    
    # Determine deductibility
    category = transaction.category
    if is_international and transaction.transaction_type == "expense":
        category = "viajes_internacionales"
    
    is_deductible = SRI_CATEGORIES.get(category, {}).get("deductible", False) and not is_international
    
    doc = {
        "id": transaction_id,
        "user_id": user["id"],
        **transaction.model_dump(),
        "category": category,
        "is_international": is_international,
        "is_deductible": is_deductible,
        "status": initial_status,
        "duplicate_of": duplicate_of,
        "match_confidence": match_confidence,
        "ai_classified": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.transactions.insert_one(doc)
    
    response_doc = {k: v for k, v in doc.items() if k != "_id"}
    
    # Add duplicate warning to response
    if duplicates:
        response_doc["_duplicate_warning"] = {
            "message": "Posible duplicado detectado",
            "potential_match": duplicates[0]["transaction"]["id"],
            "confidence": match_confidence
        }
    
    return TransactionResponse(**response_doc)

@api_router.get("/transactions", response_model=List[TransactionResponse])
async def get_transactions(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    category: Optional[str] = None,
    transaction_type: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    query = {"user_id": user["id"]}
    
    if start_date:
        query["date"] = {"$gte": start_date}
    if end_date:
        query.setdefault("date", {})["$lte"] = end_date
    if category:
        query["category"] = category
    if transaction_type:
        query["transaction_type"] = transaction_type
    
    transactions = await db.transactions.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    return [TransactionResponse(**t) for t in transactions]

@api_router.delete("/transactions/{transaction_id}")
async def delete_transaction(
    transaction_id: str,
    user: dict = Depends(get_current_user)
):
    result = await db.transactions.delete_one({"id": transaction_id, "user_id": user["id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Transacción no encontrada")
    return {"message": "Transacción eliminada"}

@api_router.put("/transactions/{transaction_id}", response_model=TransactionResponse)
async def update_transaction(
    transaction_id: str,
    transaction: TransactionCreate,
    user: dict = Depends(get_current_user)
):
    existing = await db.transactions.find_one({"id": transaction_id, "user_id": user["id"]}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Transacción no encontrada")
    
    update_data = transaction.model_dump()
    await db.transactions.update_one(
        {"id": transaction_id},
        {"$set": update_data}
    )
    
    updated = await db.transactions.find_one({"id": transaction_id}, {"_id": 0})
    return TransactionResponse(**updated)

# ================= DASHBOARD ENDPOINTS =================

@api_router.get("/dashboard/stats", response_model=DashboardStats)
async def get_dashboard_stats(user: dict = Depends(get_current_user)):
    now = datetime.now(timezone.utc)
    start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    start_of_week = now - timedelta(days=now.weekday())
    
    # Get all transactions for current month
    transactions = await db.transactions.find(
        {"user_id": user["id"], "date": {"$gte": start_of_month.strftime("%Y-%m-%d")}},
        {"_id": 0}
    ).to_list(1000)
    
    total_income = sum(t["amount"] for t in transactions if t["transaction_type"] == "income")
    total_expenses = sum(t["amount"] for t in transactions if t["transaction_type"] == "expense")
    
    # Weekly total
    weekly_transactions = [t for t in transactions if t["date"] >= start_of_week.strftime("%Y-%m-%d")]
    weekly_total = sum(t["amount"] for t in weekly_transactions if t["transaction_type"] == "expense")
    
    # Daily average
    days_in_month = now.day
    daily_average = total_expenses / days_in_month if days_in_month > 0 else 0
    
    # By category
    by_category = {}
    for t in transactions:
        if t["transaction_type"] == "expense":
            cat = t["category"]
            by_category[cat] = by_category.get(cat, 0) + t["amount"]
    
    # SRI deductible (alimentación, salud, educación, vivienda, vestimenta)
    sri_deductible_cats = ["alimentacion", "salud", "educacion", "vivienda", "vestimenta"]
    sri_deductible = sum(by_category.get(cat, 0) for cat in sri_deductible_cats)
    
    return DashboardStats(
        total_income=total_income,
        total_expenses=total_expenses,
        balance=total_income - total_expenses,
        daily_average=round(daily_average, 2),
        weekly_total=weekly_total,
        monthly_total=total_expenses,
        by_category=by_category,
        sri_deductible=sri_deductible
    )

@api_router.get("/dashboard/chart-data")
async def get_chart_data(
    period: str = "month",  # week, month, year
    user: dict = Depends(get_current_user)
):
    now = datetime.now(timezone.utc)
    
    if period == "week":
        start_date = (now - timedelta(days=7)).strftime("%Y-%m-%d")
    elif period == "month":
        start_date = now.replace(day=1).strftime("%Y-%m-%d")
    else:  # year
        start_date = now.replace(month=1, day=1).strftime("%Y-%m-%d")
    
    transactions = await db.transactions.find(
        {"user_id": user["id"], "date": {"$gte": start_date}},
        {"_id": 0}
    ).sort("date", 1).to_list(1000)
    
    # Group by date
    daily_data = {}
    for t in transactions:
        date = t["date"]
        if date not in daily_data:
            daily_data[date] = {"date": date, "income": 0, "expenses": 0}
        
        if t["transaction_type"] == "income":
            daily_data[date]["income"] += t["amount"]
        else:
            daily_data[date]["expenses"] += t["amount"]
    
    return {"data": list(daily_data.values())}

# ================= AI PROCESSING ENDPOINTS =================

async def classify_with_ai(text: str, context: str = "expense") -> dict:
    """Use OpenAI to classify transaction"""
    if not EMERGENT_LLM_KEY:
        return {"category": "otros", "subcategory": "Varios", "amount": 0, "description": text}
    
    try:
        chat = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=f"classify_{uuid.uuid4()}",
            system_message="""Eres un asistente financiero experto en leyes tributarias de Ecuador.
            Clasifica las transacciones en las siguientes categorías del SRI:
            - alimentacion: Comida, Restaurantes, Supermercado
            - salud: Seguros, Medicina, Consultas
            - educacion: Colegio y actividades, Cursos, Materiales
            - vivienda: Servicios básicos, Arriendo, Mantenimiento
            - vestimenta: Ropa, Calzado, Accesorios
            - transporte: Carros, Combustible, Mantenimiento vehicular
            - otros: Empleados, Viajes y Entretenimiento, Varios
            
            Responde SOLO en formato JSON: {"category": "...", "subcategory": "...", "amount": numero, "description": "...", "establishment": "..."}"""
        ).with_model("openai", "gpt-5.2")
        
        response = await chat.send_message(UserMessage(text=f"Clasifica esta transacción: {text}"))
        
        # Parse JSON from response
        json_match = re.search(r'\{[^}]+\}', response)
        if json_match:
            return json.loads(json_match.group())
        return {"category": "otros", "subcategory": "Varios", "amount": 0, "description": text}
    except Exception as e:
        logger.error(f"AI classification error: {e}")
        return {"category": "otros", "subcategory": "Varios", "amount": 0, "description": text}

async def process_image_with_ai(file_path: str, document_type: str = "receipt") -> dict:
    """Use Gemini to extract data from receipt/statement image or PDF"""
    if not EMERGENT_LLM_KEY:
        return {"error": "API key not configured"}
    
    try:
        ext = os.path.splitext(file_path)[1].lower()
        
        # For PDFs, try to extract text first using pdfplumber
        if ext == ".pdf" and document_type == "bank_statement":
            logger.info(f"Processing bank statement PDF: {file_path}")
            extracted_text = extract_text_from_pdf(file_path)
            
            # Only use text-based processing if we got substantial text
            if extracted_text and len(extracted_text) > 500:
                logger.info(f"Using text-based processing. Extracted {len(extracted_text)} characters")
                return await process_bank_statement_text(extracted_text)
            else:
                logger.info("PDF appears to be image-based (scanned). Using Gemini vision for OCR.")
        
        # Determine mime type for image/fallback processing
        if ext == ".pdf":
            mime_type = "application/pdf"
        elif ext in [".png"]:
            mime_type = "image/png"
        elif ext in [".jpg", ".jpeg"]:
            mime_type = "image/jpeg"
        else:
            mime_type = "image/jpeg"
        
        # Enhanced prompt specifically for Pacificard and Ecuadorian bank statements
        if document_type == "bank_statement":
            system_prompt = """Eres un experto en OCR y análisis de estados de cuenta bancarios de Ecuador.
Este documento puede ser un PDF escaneado o imagen de un estado de cuenta de Pacificard, Pichincha, Diners, Produbanco o Guayaquil.

EXTRAE TODA la información del estado de cuenta en formato JSON VÁLIDO.

INSTRUCCIONES ESPECÍFICAS PARA PACIFICARD:
- "Deuda Total a la fecha corte" o "Valor Monetario USS" = current_balance
- "Total a pagar de contado" o "PAGO SUGERIDO" = suggested_payment
- "Mínimo a pagar" = minimum_payment
- "Fecha máximo de pago sin recargos" = due_date
- "Cupo Autorizado" = credit_limit
- "Disponible" = available_credit
- "Tasa de interés efectiva anual" = apr
- "Saldo Diferido Actual" = deferred_balance
- "Saldo Rotativo Actual" = rotative_balance
- Buscar número de tarjeta formato XXXX-XXXX-XXXX-XXXX o últimos 4 dígitos
- Buscar "BLACK", "PLATINUM", "GOLD" para tipo de tarjeta

FORMATO JSON REQUERIDO:
{
  "card_info": {
    "bank_name": "Pacificard" (o banco detectado),
    "card_name": "Black/Platinum/Gold/etc",
    "card_number_last4": "últimos 4 dígitos",
    "statement_date": "YYYY-MM-DD",
    "due_date": "YYYY-MM-DD",
    "credit_limit": numero,
    "available_credit": numero,
    "current_balance": numero,
    "minimum_payment": numero,
    "suggested_payment": numero,
    "apr": numero,
    "previous_balance": numero,
    "deferred_balance": numero,
    "rotative_balance": numero
  },
  "transactions": [
    {
      "date": "YYYY-MM-DD",
      "description": "descripción del consumo",
      "amount": numero positivo,
      "establishment": "nombre comercio",
      "category": "alimentacion|salud|educacion|vivienda|vestimenta|suscripciones|turismo|transporte|entretenimiento|otros",
      "is_international": false,
      "is_subscription": false,
      "is_fee": false
    }
  ],
  "deferred_purchases": [
    {
      "description": "descripción de compra diferida",
      "total_amount": numero,
      "monthly_payment": numero,
      "current_installment": numero,
      "remaining_installments": numero,
      "total_installments": numero
    }
  ]
}

CATEGORIZACIÓN AUTOMÁTICA:
- SUPERMAXI, MEGAMAXI, KORTES, RED CRAB, DUNKIN, CASADELI, CIABATTA, MIGAJAS, ROGERS, LA COSTENITA, NELSON MARKET = alimentacion
- DISNEY PLUS, APPLE.COM/BILL, NETFLIX, SPOTIFY, HBO, YOUTUBE PREMIUM = suscripciones (is_subscription: true)
- DIFARE, COMFARPI, FYBECA, MEDICITY = salud
- ESTACION DE SERVICIO, PRIMAX, PDV, TEXACO = transporte
- ZARA, H&M, ETAFASHION, TENIS MARKET, FOREVERSOFT = vestimenta
- SUPERCINES, RIVER BEACH TENNIS, DIAMOND CLUB = entretenimiento
- CONECEL, CLARO, MOVISTAR, CNT = vivienda
- AMAGUA, CISNERGIA, CNEL, LUZ = vivienda
- GESTION DE COBRANZA, INTERES, CONTRIB.FINANC.SOLCA, COMISION = is_fee: true
- Consumos marcados como EXTERIOR = is_international: true
- IVA SERVICIO DIGITAL = is_international: true, is_fee: false

REGLAS IMPORTANTES:
- Convertir fechas: ENE=01, FEB=02, MAR=03, ABR=04, MAY=05, JUN=06, JUL=07, AGO=08, SEP=09, OCT=10, NOV=11, DIC=12
- El año actual es 2025/2026
- NO incluir pagos (líneas con "SU PAGO" o montos negativos)
- Para diferidos tipo "PACIFICARD EFECTIVO BANCA" o marcados como "DIF", extraer cuota mensual
- Buscar sección "DETALLE DE MOVIMIENTOS DEL PERÍODO" para transacciones
- Cada transacción debe tener amount > 0"""
            
            user_prompt = """Analiza este estado de cuenta bancario ecuatoriano (puede ser Pacificard, Pichincha u otro banco).
Extrae TODA la información: datos de la tarjeta, TODAS las transacciones del período, y compras diferidas.
Responde ÚNICAMENTE con JSON válido, sin explicaciones adicionales."""
        else:
            system_prompt = """Eres un experto en OCR y análisis de recibos/facturas ecuatorianas.
            Extrae la información del recibo y clasifícala según categorías SRI Ecuador.
            Responde SOLO en formato JSON: 
            {"transactions": [{"amount": numero, "description": "...", "category": "...", "subcategory": "...", "establishment": "...", "date": "YYYY-MM-DD"}]}"""
            user_prompt = "Extrae toda la información de este recibo/factura ecuatoriana"
        
        logger.info(f"Sending file to Gemini for OCR processing: {file_path}")
        
        chat = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=f"ocr_{uuid.uuid4()}",
            system_message=system_prompt
        ).with_model("gemini", "gemini-2.5-flash")
        
        file_content = FileContentWithMimeType(
            file_path=file_path,
            mime_type=mime_type
        )
        
        response = await chat.send_message(UserMessage(
            text=user_prompt,
            file_contents=[file_content]
        ))
        
        logger.info(f"Gemini response received, length: {len(response)}")
        
        # Parse JSON from response
        json_match = re.search(r'\{[\s\S]*\}', response)
        if json_match:
            try:
                result = json.loads(json_match.group())
                logger.info(f"Successfully parsed JSON - card_info: {bool(result.get('card_info'))}, transactions: {len(result.get('transactions', []))}")
                return result
            except json.JSONDecodeError as e:
                logger.error(f"JSON decode error: {e}")
                # Try to fix common JSON issues
                try:
                    fixed = json_match.group().replace("'", '"')
                    return json.loads(fixed)
                except:
                    pass
        
        logger.warning("No valid JSON found in Gemini response")
        return {"transactions": [], "card_info": None}
    except Exception as e:
        logger.error(f"Image processing error: {e}")
        return {"transactions": [], "card_info": None, "error": str(e)}


def extract_text_from_pdf(file_path: str) -> str:
    """Extract all text from PDF using pdfplumber for better accuracy"""
    try:
        all_text = []
        with pdfplumber.open(file_path) as pdf:
            for page_num, page in enumerate(pdf.pages):
                # Extract text
                text = page.extract_text()
                if text:
                    all_text.append(f"--- PÁGINA {page_num + 1} ---\n{text}")
                
                # Also try to extract tables which are common in bank statements
                tables = page.extract_tables()
                for table in tables:
                    if table:
                        table_text = "\n".join([" | ".join([str(cell or "") for cell in row]) for row in table if row])
                        if table_text.strip():
                            all_text.append(f"\n[TABLA]\n{table_text}")
        
        return "\n\n".join(all_text)
    except Exception as e:
        logger.error(f"PDF text extraction error: {e}")
        return ""


async def process_bank_statement_text(text: str) -> dict:
    """Process extracted bank statement text with AI"""
    if not EMERGENT_LLM_KEY:
        return {"error": "API key not configured"}
    
    try:
        system_prompt = """Eres un experto en análisis de estados de cuenta bancarios de Ecuador.
Analiza el texto extraído del estado de cuenta y extrae TODA la información en formato JSON.

INSTRUCCIONES ESPECÍFICAS PARA PACIFICARD:
- "Deuda Total a la fecha corte" = current_balance (saldo total a pagar)
- "Total a pagar de contado" o "PAGO SUGERIDO" = pago de contado recomendado
- "Mínimo a pagar" o "Pago Mínimo" = minimum_payment
- "Fecha máximo de pago sin recargos" = due_date
- "Cupo Autorizado" = credit_limit
- "Disponible" = available_credit
- "Tasa de interés efectiva anual" = apr
- "Saldo Diferido Actual" = total de compras diferidas pendientes
- Buscar "BLACK" o "PLATINUM" en el nombre de la tarjeta

FORMATO DE RESPUESTA JSON:
{
  "card_info": {
    "bank_name": "Pacificard" o el banco detectado,
    "card_name": "Black/Platinum/etc",
    "card_number_last4": "últimos 4 dígitos si se encuentran",
    "statement_date": "YYYY-MM-DD",
    "due_date": "YYYY-MM-DD",
    "credit_limit": numero,
    "available_credit": numero,
    "current_balance": numero (deuda total),
    "minimum_payment": numero,
    "apr": numero (tasa efectiva anual),
    "previous_balance": numero,
    "payments_received": numero,
    "period_charges": numero,
    "deferred_balance": numero (saldo diferido)
  },
  "transactions": [
    {
      "date": "YYYY-MM-DD",
      "description": "descripción",
      "amount": numero positivo,
      "establishment": "comercio",
      "category": "alimentacion|salud|educacion|vivienda|vestimenta|suscripciones|turismo|transporte|entretenimiento|impuestos|otros",
      "is_international": boolean,
      "is_subscription": boolean,
      "is_fee": boolean
    }
  ],
  "deferred_purchases": [
    {
      "description": "descripción",
      "total_amount": numero,
      "monthly_payment": numero (cuota mensual),
      "current_installment": numero (cuota actual X de Y),
      "remaining_installments": numero,
      "total_installments": numero
    }
  ]
}

CATEGORIZACIÓN:
- DIAMOND CLUB, SUPERMAXI, KORTES, MEGAMAXI, RED CRAB, DUNKIN, CASADELI, CIABATTA, MIGAJAS = alimentacion
- DISNEY PLUS, APPLE.COM/BILL, NETFLIX, SPOTIFY, HBO = suscripciones (is_subscription: true)
- DIFARE, COMFARPI, FYBECA = salud
- ESTACION DE SERVICIO, PRIMAX, PDV = transporte
- ZARA, KORTES SOL PLAZA = vestimenta
- SUPERCINES, RIVER BEACH TENNIS = entretenimiento
- CONECEL, CLARO = vivienda (telefonía)
- AMAGUA, CISNERGIA = vivienda (servicios básicos)
- GESTION DE COBRANZA, INTERES, CONTRIB.FINANC = is_fee: true
- Consumos con "EXTERIOR" o moneda extranjera = is_international: true

REGLAS:
- Las fechas DD/MMM/AAAA convertir a YYYY-MM-DD (ENE=01, FEB=02, MAR=03, ABR=04, MAY=05, JUN=06, JUL=07, AGO=08, SEP=09, OCT=10, NOV=11, DIC=12)
- Incluir TODAS las transacciones del período
- NO incluir pagos (montos negativos o "SU PAGO")
- Para diferidos, buscar sección "DIF" o cuotas mensuales recurrentes"""

        user_prompt = f"""Analiza este texto extraído del estado de cuenta bancario y extrae TODA la información:

{text[:15000]}

Responde SOLO con el JSON estructurado."""

        chat = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=f"bank_stmt_{uuid.uuid4()}",
            system_message=system_prompt
        ).with_model("gemini", "gemini-2.5-flash")
        
        response = await chat.send_message(UserMessage(text=user_prompt))
        
        logger.info(f"AI response length: {len(response)}")
        
        # Parse JSON from response
        json_match = re.search(r'\{[\s\S]*\}', response)
        if json_match:
            try:
                result = json.loads(json_match.group())
                logger.info(f"Parsed result - card_info: {bool(result.get('card_info'))}, transactions: {len(result.get('transactions', []))}, deferred: {len(result.get('deferred_purchases', []))}")
                return result
            except json.JSONDecodeError as e:
                logger.error(f"JSON decode error: {e}")
                # Try to fix common JSON issues
                try:
                    fixed = json_match.group().replace("'", '"')
                    return json.loads(fixed)
                except:
                    pass
        
        logger.warning("No valid JSON found in AI response")
        return {"transactions": [], "card_info": None}
    except Exception as e:
        logger.error(f"Bank statement text processing error: {e}")
        return {"transactions": [], "card_info": None, "error": str(e)}

@api_router.post("/process/email")
async def process_email(
    email_content: str = Form(...),
    user: dict = Depends(get_current_user)
):
    """Process credit card email notification"""
    # Parse PacifiCard format
    result = await classify_with_ai(email_content, "email")
    
    # Create transaction
    transaction_id = str(uuid.uuid4())
    doc = {
        "id": transaction_id,
        "user_id": user["id"],
        "amount": result.get("amount", 0),
        "description": result.get("description", ""),
        "category": result.get("category", "otros"),
        "subcategory": result.get("subcategory", "Varios"),
        "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "transaction_type": "expense",
        "establishment": result.get("establishment", ""),
        "ai_classified": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.transactions.insert_one(doc)
    
    return {"message": "Email procesado", "transaction": {k: v for k, v in doc.items() if k != "_id"}}

@api_router.post("/process/receipt")
async def process_receipt(
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user)
):
    """Process receipt image with OCR"""
    # Save file temporarily
    temp_dir = tempfile.mkdtemp()
    file_path = os.path.join(temp_dir, file.filename)
    
    async with aiofiles.open(file_path, 'wb') as f:
        content = await file.read()
        await f.write(content)
    
    try:
        result = await process_image_with_ai(file_path)
        
        created_transactions = []
        for t in result.get("transactions", []):
            transaction_id = str(uuid.uuid4())
            doc = {
                "id": transaction_id,
                "user_id": user["id"],
                "amount": t.get("amount", 0),
                "description": t.get("description", ""),
                "category": t.get("category", "otros"),
                "subcategory": t.get("subcategory", "Varios"),
                "date": t.get("date", datetime.now(timezone.utc).strftime("%Y-%m-%d")),
                "transaction_type": "expense",
                "establishment": t.get("establishment", ""),
                "ai_classified": True,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.transactions.insert_one(doc)
            created_transactions.append({k: v for k, v in doc.items() if k != "_id"})
        
        return {"message": f"Procesadas {len(created_transactions)} transacciones", "transactions": created_transactions}
    finally:
        # Cleanup
        if os.path.exists(file_path):
            os.remove(file_path)
        os.rmdir(temp_dir)

@api_router.post("/process/receipts-multiple")
async def process_multiple_receipts(
    files: List[UploadFile] = File(...),
    user: dict = Depends(get_current_user)
):
    """Process multiple receipt images with OCR"""
    all_transactions = []
    errors = []
    
    for file in files:
        temp_dir = tempfile.mkdtemp()
        file_path = os.path.join(temp_dir, file.filename)
        
        try:
            async with aiofiles.open(file_path, 'wb') as f:
                content = await file.read()
                await f.write(content)
            
            result = await process_image_with_ai(file_path)
            
            for t in result.get("transactions", []):
                # Check if international
                country = t.get("country", "Ecuador")
                is_international = any(c.lower() in country.lower() for c in INTERNATIONAL_COUNTRIES) if country else False
                category = "viajes_internacionales" if is_international else t.get("category", "otros")
                is_deductible = SRI_CATEGORIES.get(category, {}).get("deductible", False)
                
                amount = t.get("amount", 0)
                date = t.get("date", datetime.now(timezone.utc).strftime("%Y-%m-%d"))
                establishment = t.get("establishment", "")
                description = t.get("description", "")
                
                # Check for duplicates
                duplicates = await find_potential_duplicates(user["id"], amount, date, establishment, description)
                
                transaction_id = str(uuid.uuid4())
                doc = {
                    "id": transaction_id,
                    "user_id": user["id"],
                    "amount": amount,
                    "description": description,
                    "category": category,
                    "subcategory": t.get("subcategory", "Varios"),
                    "date": date,
                    "transaction_type": "expense",
                    "establishment": establishment,
                    "country": country,
                    "is_international": is_international,
                    "payment_source": "internacional" if is_international else "local",
                    "is_deductible": is_deductible,
                    "ai_classified": True,
                    "status": TransactionStatus.DUPLICATE_SUSPECT if duplicates else TransactionStatus.PENDING_REVIEW,
                    "source_type": SourceType.RECEIPT,
                    "has_receipt": True,
                    "duplicate_of": duplicates[0]["transaction"]["id"] if duplicates else None,
                    "match_confidence": duplicates[0]["confidence"] if duplicates else None,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "source_file": file.filename
                }
                await db.transactions.insert_one(doc)
                all_transactions.append({k: v for k, v in doc.items() if k != "_id"})
        except Exception as e:
            errors.append({"file": file.filename, "error": str(e)})
        finally:
            if os.path.exists(file_path):
                os.remove(file_path)
            if os.path.exists(temp_dir):
                os.rmdir(temp_dir)
    
    return {
        "message": f"Procesados {len(files)} archivos, {len(all_transactions)} transacciones creadas",
        "transactions": all_transactions,
        "errors": errors
    }

@api_router.post("/process/bank-statement")
async def process_bank_statement(
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user)
):
    """
    Process bank/credit card statement PDF or image.
    Extracts: card info (balance, minimum payment, APR, dates) AND all transactions.
    """
    temp_dir = tempfile.mkdtemp()
    file_path = os.path.join(temp_dir, file.filename)
    
    try:
        async with aiofiles.open(file_path, 'wb') as f:
            content = await file.read()
            await f.write(content)
        
        # Process with AI specifically for bank statements
        result = await process_image_with_ai(file_path, document_type="bank_statement")
        
        response_data = {
            "card_info": None,
            "card_updated": False,
            "transactions_created": 0,
            "transactions": []
        }
        
        # Process card info if found
        card_info = result.get("card_info")
        if card_info and card_info.get("current_balance"):
            # Try to find existing card or create new one
            bank_name = card_info.get("bank_name", "").lower()
            card_number = card_info.get("card_number_last4", "")
            
            # Search for existing card
            existing_card = None
            if bank_name:
                existing_card = await db.credit_cards.find_one({
                    "user_id": user["id"],
                    "$or": [
                        {"name": {"$regex": bank_name, "$options": "i"}},
                        {"last_four_digits": card_number}
                    ]
                })
            
            if existing_card:
                # Update existing card with new info
                update_data = {
                    "current_balance": card_info.get("current_balance", existing_card.get("current_balance", 0)),
                    "minimum_payment": card_info.get("minimum_payment", existing_card.get("minimum_payment", 0)),
                    "credit_limit": card_info.get("credit_limit", existing_card.get("credit_limit", 0)),
                    "available_credit": card_info.get("available_credit"),
                    "statement_date": card_info.get("statement_date"),
                    "due_date": card_info.get("due_date"),
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
                if card_info.get("apr"):
                    update_data["apr"] = card_info["apr"]
                
                await db.credit_cards.update_one(
                    {"id": existing_card["id"]},
                    {"$set": update_data}
                )
                response_data["card_updated"] = True
                response_data["card_info"] = {**existing_card, **update_data, "_id": None}
                
                # Create scheduled payment in flow for card payment
                if card_info.get("due_date") and card_info.get("current_balance"):
                    due_date = card_info["due_date"]
                    try:
                        due_day = int(due_date.split("-")[2]) if "-" in due_date else 15
                    except:
                        due_day = 15
                    
                    # Check if payment already exists for this card this month
                    existing_payment = await db.scheduled_payments.find_one({
                        "user_id": user["id"],
                        "card_id": existing_card["id"],
                        "month": datetime.now().month
                    })
                    
                    if not existing_payment:
                        payment_doc = {
                            "id": str(uuid.uuid4()),
                            "user_id": user["id"],
                            "card_id": existing_card["id"],
                            "description": f"Pago Tarjeta {existing_card['name']}",
                            "amount": card_info.get("current_balance", 0),
                            "minimum_amount": card_info.get("minimum_payment", 0),
                            "due_day": due_day,
                            "due_date": due_date,
                            "month": datetime.now().month,
                            "year": datetime.now().year,
                            "payment_method": "transferencia",
                            "category": "tarjeta_credito",
                            "is_recurring": True,
                            "is_card_payment": True,
                            "status": "pending",
                            "created_at": datetime.now(timezone.utc).isoformat()
                        }
                        await db.scheduled_payments.insert_one(payment_doc)
                        response_data["payment_scheduled"] = True
            else:
                # Create new card
                new_card_id = str(uuid.uuid4())
                new_card = {
                    "id": new_card_id,
                    "user_id": user["id"],
                    "name": card_info.get("card_name") or card_info.get("bank_name", "Tarjeta Importada"),
                    "bank": card_info.get("bank_name", ""),
                    "last_four_digits": card_number,
                    "credit_limit": card_info.get("credit_limit", 0),
                    "current_balance": card_info.get("current_balance", 0),
                    "minimum_payment": card_info.get("minimum_payment", 0),
                    "apr": card_info.get("apr", 0),
                    "statement_date": card_info.get("statement_date"),
                    "due_date": card_info.get("due_date"),
                    "available_credit": card_info.get("available_credit"),
                    "currency": "USD",
                    "is_international": False,
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.credit_cards.insert_one(new_card)
                response_data["card_info"] = {k: v for k, v in new_card.items() if k != "_id"}
                response_data["card_updated"] = True
                
                # Create scheduled payment for new card
                if card_info.get("due_date") and card_info.get("current_balance"):
                    due_date = card_info["due_date"]
                    try:
                        due_day = int(due_date.split("-")[2]) if "-" in due_date else 15
                    except:
                        due_day = 15
                    
                    payment_doc = {
                        "id": str(uuid.uuid4()),
                        "user_id": user["id"],
                        "card_id": new_card_id,
                        "description": f"Pago Tarjeta {new_card['name']}",
                        "amount": card_info.get("current_balance", 0),
                        "minimum_amount": card_info.get("minimum_payment", 0),
                        "due_day": due_day,
                        "due_date": due_date,
                        "month": datetime.now().month,
                        "year": datetime.now().year,
                        "payment_method": "transferencia",
                        "category": "tarjeta_credito",
                        "is_recurring": True,
                        "is_card_payment": True,
                        "status": "pending",
                        "created_at": datetime.now(timezone.utc).isoformat()
                    }
                    await db.scheduled_payments.insert_one(payment_doc)
                    response_data["payment_scheduled"] = True
        
        # Process deferred purchases (diferidos)
        deferred_purchases = result.get("deferred_purchases", [])
        deferred_created = []
        for dp in deferred_purchases:
            if dp.get("remaining_installments", 0) > 0:
                deferred_doc = {
                    "id": str(uuid.uuid4()),
                    "user_id": user["id"],
                    "description": dp.get("description", "Compra Diferida"),
                    "total_amount": dp.get("total_amount", 0),
                    "monthly_payment": dp.get("monthly_payment", 0),
                    "remaining_installments": dp.get("remaining_installments", 0),
                    "total_installments": dp.get("total_installments", dp.get("remaining_installments", 0)),
                    "card_id": response_data.get("card_info", {}).get("id"),
                    "card_name": card_info.get("card_name") or card_info.get("bank_name") if card_info else None,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "source_file": file.filename
                }
                await db.deferred_payments.insert_one(deferred_doc)
                deferred_created.append({k: v for k, v in deferred_doc.items() if k != "_id"})
        
        response_data["deferred_payments_created"] = len(deferred_created)
        response_data["deferred_payments"] = deferred_created
        
        # Process transactions
        transactions = result.get("transactions", [])
        created_transactions = []
        
        for t in transactions:
            amount = abs(t.get("amount", 0))  # Always positive for expenses
            if amount == 0:
                continue
                
            is_payment = t.get("amount", 0) < 0 or "pago" in t.get("description", "").lower()
            
            if is_payment:
                # This is a payment to the card, skip or handle differently
                continue
            
            date = t.get("date", datetime.now(timezone.utc).strftime("%Y-%m-%d"))
            establishment = t.get("establishment", t.get("description", "")[:50])
            description = t.get("description", "")
            
            # Detect subscriptions
            is_subscription = t.get("is_subscription", False)
            desc_lower = description.lower()
            estab_lower = establishment.lower() if establishment else ""
            
            for sub in SUBSCRIPTION_SERVICES:
                if sub in desc_lower or sub in estab_lower:
                    is_subscription = True
                    break
            
            # Override category for subscriptions - they are local, not international
            category = t.get("category", "otros")
            if is_subscription:
                category = "suscripciones"
            
            # Check for duplicates
            duplicates = await find_potential_duplicates(user["id"], amount, date, establishment, description)
            
            transaction_id = str(uuid.uuid4())
            doc = {
                "id": transaction_id,
                "user_id": user["id"],
                "amount": amount,
                "description": description,
                "category": category,
                "subcategory": t.get("subcategory", "Varios"),
                "date": date,
                "transaction_type": "expense",
                "establishment": establishment,
                "is_international": t.get("is_international", False) if not is_subscription else False,
                "is_subscription": is_subscription,
                "is_recurring": is_subscription,
                "tags": ["recurrente", "suscripcion"] if is_subscription else [],
                "payment_method": "tarjeta",
                "card_name": card_info.get("card_name") or card_info.get("bank_name") if card_info else None,
                "ai_classified": True,
                "status": TransactionStatus.DUPLICATE_SUSPECT if duplicates else TransactionStatus.PENDING_REVIEW,
                "source_type": SourceType.BANK_STATEMENT,
                "duplicate_of": duplicates[0]["transaction"]["id"] if duplicates else None,
                "match_confidence": duplicates[0]["confidence"] if duplicates else None,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "source_file": file.filename
            }
            await db.transactions.insert_one(doc)
            created_transactions.append({k: v for k, v in doc.items() if k != "_id"})
        
        response_data["transactions_created"] = len(created_transactions)
        response_data["transactions"] = created_transactions
        
        return {
            "message": f"Estado de cuenta procesado: {len(created_transactions)} transacciones, {len(deferred_created)} diferidos",
            "data": response_data,
            "raw_extraction": result if not created_transactions else None
        }
        
    except Exception as e:
        logger.error(f"Bank statement processing error: {e}")
        raise HTTPException(status_code=500, detail=f"Error procesando estado de cuenta: {str(e)}")
    finally:
        if os.path.exists(file_path):
            os.remove(file_path)
        if os.path.exists(temp_dir):
            os.rmdir(temp_dir)

@api_router.get("/transactions/international")
async def get_international_transactions(
    user: dict = Depends(get_current_user)
):
    """Get all international transactions"""
    transactions = await db.transactions.find(
        {"user_id": user["id"], "is_international": True},
        {"_id": 0}
    ).sort("date", -1).to_list(1000)
    return {"transactions": transactions}

@api_router.get("/transactions/by-payment-source")
async def get_transactions_by_payment_source(
    payment_source: str = "internacional",
    user: dict = Depends(get_current_user)
):
    """Get transactions by payment source (local or internacional)"""
    transactions = await db.transactions.find(
        {"user_id": user["id"], "payment_source": payment_source},
        {"_id": 0}
    ).sort("date", -1).to_list(1000)
    return {"transactions": transactions, "payment_source": payment_source}

@api_router.get("/budget/suggestions")
async def get_budget_suggestions(user: dict = Depends(get_current_user)):
    """Get AI-powered budget adjustment suggestions based on historical data"""
    # Get last 6 months of transactions
    now = datetime.now(timezone.utc)
    six_months_ago = (now - timedelta(days=180)).strftime("%Y-%m-%d")
    
    transactions = await db.transactions.find(
        {"user_id": user["id"], "date": {"$gte": six_months_ago}, "transaction_type": "expense"},
        {"_id": 0}
    ).to_list(5000)
    
    if len(transactions) < 10:
        return {"suggestions": [], "message": "Necesitas más transacciones para obtener sugerencias (mínimo 10)"}
    
    # Calculate monthly averages by category
    monthly_data = {}
    for t in transactions:
        month = t["date"][:7]
        cat = t["category"]
        if month not in monthly_data:
            monthly_data[month] = {}
        if cat not in monthly_data[month]:
            monthly_data[month][cat] = 0
        monthly_data[month][cat] += t["amount"]
    
    # Calculate averages and trends
    category_stats = {}
    for month, cats in monthly_data.items():
        for cat, amount in cats.items():
            if cat not in category_stats:
                category_stats[cat] = {"amounts": [], "total": 0}
            category_stats[cat]["amounts"].append(amount)
            category_stats[cat]["total"] += amount
    
    # Get current budget
    budget = await db.budgets.find_one({"user_id": user["id"]}, {"_id": 0}, sort=[("created_at", -1)])
    budget_by_cat = {}
    if budget:
        for item in budget.get("items", []):
            cat = item["category"]
            budget_by_cat[cat] = budget_by_cat.get(cat, 0) + item["planned_amount"]
    
    suggestions = []
    for cat, stats in category_stats.items():
        avg = sum(stats["amounts"]) / len(stats["amounts"])
        current_budget = budget_by_cat.get(cat, 0)
        cat_info = SRI_CATEGORIES.get(cat, {"name": cat, "deductible": False})
        
        # Check if consistently over or under budget
        if current_budget > 0:
            ratio = avg / current_budget
            if ratio > 1.2:  # 20% over budget consistently
                suggestions.append({
                    "category": cat,
                    "category_name": cat_info["name"],
                    "type": "increase",
                    "current_budget": current_budget,
                    "suggested_budget": round(avg * 1.1, 2),
                    "average_spending": round(avg, 2),
                    "reason": f"Gastas en promedio ${avg:.2f}/mes, 20% más que tu presupuesto de ${current_budget:.2f}",
                    "is_deductible": cat_info.get("deductible", False)
                })
            elif ratio < 0.6:  # 40% under budget consistently
                suggestions.append({
                    "category": cat,
                    "category_name": cat_info["name"],
                    "type": "decrease",
                    "current_budget": current_budget,
                    "suggested_budget": round(avg * 1.2, 2),
                    "average_spending": round(avg, 2),
                    "reason": f"Gastas en promedio ${avg:.2f}/mes, mucho menos que tu presupuesto de ${current_budget:.2f}",
                    "is_deductible": cat_info.get("deductible", False)
                })
        elif avg > 50:  # No budget set but significant spending
            suggestions.append({
                "category": cat,
                "category_name": cat_info["name"],
                "type": "new",
                "current_budget": 0,
                "suggested_budget": round(avg * 1.1, 2),
                "average_spending": round(avg, 2),
                "reason": f"No tienes presupuesto para {cat_info['name']} pero gastas ${avg:.2f}/mes en promedio",
                "is_deductible": cat_info.get("deductible", False)
            })
    
    # Sort by impact (difference between current and suggested)
    suggestions.sort(key=lambda x: abs(x["suggested_budget"] - x["current_budget"]), reverse=True)
    
    return {"suggestions": suggestions[:10], "months_analyzed": len(monthly_data)}

@api_router.post("/process/excel")
async def process_excel(
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user)
):
    """Process Excel budget file"""
    content = await file.read()
    
    try:
        wb = openpyxl.load_workbook(BytesIO(content))
        
        budget_items = []
        total_income = 0
        total_expenses = 0
        
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] and isinstance(row[0], str):
                    # Try to parse as budget item
                    category_name = str(row[0]).lower()
                    
                    # Map to SRI categories
                    category_map = {
                        "servicios basicos": ("vivienda", "Servicios básicos"),
                        "empleados": ("otros", "Empleados"),
                        "colegio": ("educacion", "Colegio y actividades"),
                        "seguros": ("salud", "Seguros"),
                        "comida": ("alimentacion", "Comida"),
                        "restaurantes": ("alimentacion", "Restaurantes"),
                        "carros": ("transporte", "Carros"),
                        "viajes": ("otros", "Viajes y Entretenimiento"),
                    }
                    
                    for key, (cat, subcat) in category_map.items():
                        if key in category_name:
                            amounts = [v for v in row[1:] if isinstance(v, (int, float))]
                            if amounts:
                                avg_amount = sum(amounts) / len(amounts)
                                budget_items.append({
                                    "category": cat,
                                    "subcategory": subcat,
                                    "planned_amount": avg_amount,
                                    "month": datetime.now(timezone.utc).strftime("%Y-%m")
                                })
                                total_expenses += avg_amount
                            break
                    
                    # Check for income
                    if "ingreso" in category_name or "personal" in category_name or "apx" in category_name:
                        amounts = [v for v in row[1:] if isinstance(v, (int, float))]
                        if amounts:
                            total_income += sum(amounts) / len(amounts)
        
        # Save budget
        budget_id = str(uuid.uuid4())
        budget_doc = {
            "id": budget_id,
            "user_id": user["id"],
            "items": budget_items,
            "total_income": total_income,
            "total_expenses": total_expenses,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.budgets.insert_one(budget_doc)
        
        return {"message": "Excel procesado", "budget": {k: v for k, v in budget_doc.items() if k != "_id"}}
    except Exception as e:
        logger.error(f"Excel processing error: {e}")
        raise HTTPException(status_code=400, detail=f"Error procesando Excel: {str(e)}")

# ================= BUDGET ENDPOINTS =================

@api_router.get("/budget", response_model=BudgetResponse)
async def get_budget(
    month: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    query = {"user_id": user["id"]}
    budget = await db.budgets.find_one(query, {"_id": 0}, sort=[("created_at", -1)])
    
    if not budget:
        # Return empty budget
        return BudgetResponse(
            id="",
            user_id=user["id"],
            items=[],
            total_income=0,
            total_expenses=0,
            created_at=datetime.now(timezone.utc).isoformat()
        )
    
    return BudgetResponse(**budget)

@api_router.get("/budget/vs-actual")
async def get_budget_vs_actual(user: dict = Depends(get_current_user)):
    """Compare budget vs actual spending"""
    # Get latest budget
    budget = await db.budgets.find_one({"user_id": user["id"]}, {"_id": 0}, sort=[("created_at", -1)])
    
    # Get current month transactions
    now = datetime.now(timezone.utc)
    start_of_month = now.replace(day=1).strftime("%Y-%m-%d")
    
    transactions = await db.transactions.find(
        {"user_id": user["id"], "date": {"$gte": start_of_month}, "transaction_type": "expense"},
        {"_id": 0}
    ).to_list(1000)
    
    # Group actual by category
    actual_by_category = {}
    for t in transactions:
        cat = t["category"]
        actual_by_category[cat] = actual_by_category.get(cat, 0) + t["amount"]
    
    # Compare
    comparison = []
    if budget:
        budget_by_category = {}
        for item in budget.get("items", []):
            cat = item["category"]
            budget_by_category[cat] = budget_by_category.get(cat, 0) + item["planned_amount"]
        
        for cat in set(list(budget_by_category.keys()) + list(actual_by_category.keys())):
            planned = budget_by_category.get(cat, 0)
            actual = actual_by_category.get(cat, 0)
            comparison.append({
                "category": cat,
                "category_name": SRI_CATEGORIES.get(cat, {}).get("name", cat),
                "planned": planned,
                "actual": actual,
                "difference": planned - actual,
                "percentage": (actual / planned * 100) if planned > 0 else 0
            })
    
    return {"comparison": comparison}

# ================= AI PREDICTIONS ENDPOINTS =================

@api_router.get("/predictions")
async def get_predictions(user: dict = Depends(get_current_user)):
    """Get AI-powered spending predictions and advice"""
    if not EMERGENT_LLM_KEY:
        return {"predictions": [], "advice": ["Configure API key para predicciones AI"], "sri_tips": []}
    
    # Get last 3 months of transactions
    now = datetime.now(timezone.utc)
    three_months_ago = (now - timedelta(days=90)).strftime("%Y-%m-%d")
    
    transactions = await db.transactions.find(
        {"user_id": user["id"], "date": {"$gte": three_months_ago}},
        {"_id": 0}
    ).to_list(1000)
    
    if not transactions:
        return {"predictions": [], "advice": ["Agrega transacciones para obtener predicciones"], "sri_tips": []}
    
    # Prepare data for AI
    summary = {}
    for t in transactions:
        cat = t["category"]
        if cat not in summary:
            summary[cat] = {"total": 0, "count": 0, "transactions": []}
        summary[cat]["total"] += t["amount"]
        summary[cat]["count"] += 1
    
    try:
        chat = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=f"predict_{uuid.uuid4()}",
            system_message="""Eres un asesor financiero experto en finanzas personales y leyes tributarias de Ecuador.
            Analiza los gastos y proporciona:
            1. Predicciones de gastos para el próximo mes por categoría
            2. Consejos específicos para optimizar recursos
            3. Recomendaciones para maximizar deducciones SRI
            
            Responde en formato JSON:
            {
                "predictions": [{"category": "...", "predicted_amount": numero, "trend": "up/down/stable"}],
                "advice": ["consejo1", "consejo2", ...],
                "sri_tips": ["tip1", "tip2", ...]
            }"""
        ).with_model("openai", "gpt-5.2")
        
        response = await chat.send_message(UserMessage(
            text=f"Analiza estos gastos de los últimos 3 meses y proporciona predicciones: {json.dumps(summary)}"
        ))
        
        json_match = re.search(r'\{.*\}', response, re.DOTALL)
        if json_match:
            result = json.loads(json_match.group())
            return result
        
        return {"predictions": [], "advice": ["No se pudo generar predicciones"], "sri_tips": []}
    except Exception as e:
        logger.error(f"Prediction error: {e}")
        return {"predictions": [], "advice": [f"Error: {str(e)}"], "sri_tips": []}

# ================= ACCOUNTANT VIEW ENDPOINTS =================

@api_router.get("/accountant/tax-summary")
async def get_tax_summary(
    year: Optional[int] = None,
    user: dict = Depends(check_role([UserRole.ADMIN, UserRole.ACCOUNTANT]))
):
    """Get tax deductible summary for accountant"""
    if not year:
        year = datetime.now(timezone.utc).year
    
    start_date = f"{year}-01-01"
    end_date = f"{year}-12-31"
    
    # Get all users' transactions (for accountant view)
    transactions = await db.transactions.find(
        {"date": {"$gte": start_date, "$lte": end_date}, "transaction_type": "expense"},
        {"_id": 0}
    ).to_list(10000)
    
    # SRI deductible categories
    sri_deductible_cats = ["alimentacion", "salud", "educacion", "vivienda", "vestimenta"]
    
    summary = {
        "year": year,
        "total_expenses": 0,
        "deductible_expenses": 0,
        "by_category": {},
        "monthly_breakdown": {}
    }
    
    for t in transactions:
        cat = t["category"]
        month = t["date"][:7]
        amount = t["amount"]
        
        summary["total_expenses"] += amount
        
        if cat in sri_deductible_cats:
            summary["deductible_expenses"] += amount
        
        if cat not in summary["by_category"]:
            summary["by_category"][cat] = {
                "name": SRI_CATEGORIES.get(cat, {}).get("name", cat),
                "total": 0,
                "deductible": cat in sri_deductible_cats
            }
        summary["by_category"][cat]["total"] += amount
        
        if month not in summary["monthly_breakdown"]:
            summary["monthly_breakdown"][month] = {"total": 0, "deductible": 0}
        summary["monthly_breakdown"][month]["total"] += amount
        if cat in sri_deductible_cats:
            summary["monthly_breakdown"][month]["deductible"] += amount
    
    return summary

# ================= RECONCILIATION ENDPOINTS (CONTADORA) =================

@api_router.get("/reconciliation/pending")
async def get_pending_reconciliation(
    user: dict = Depends(check_role([UserRole.ADMIN, UserRole.ACCOUNTANT]))
):
    """Get all transactions pending review - Vista Contadora"""
    # Get transactions that need review
    transactions = await db.transactions.find(
        {"status": {"$in": [TransactionStatus.PENDING_REVIEW, TransactionStatus.DUPLICATE_SUSPECT]}},
        {"_id": 0}
    ).sort("created_at", -1).to_list(500)
    
    # Group by status
    pending_review = [t for t in transactions if t.get("status") == TransactionStatus.PENDING_REVIEW]
    duplicate_suspects = [t for t in transactions if t.get("status") == TransactionStatus.DUPLICATE_SUSPECT]
    
    # Get stats
    total_pending_amount = sum(t.get("amount", 0) for t in pending_review)
    total_duplicate_amount = sum(t.get("amount", 0) for t in duplicate_suspects)
    
    return {
        "pending_review": pending_review,
        "duplicate_suspects": duplicate_suspects,
        "stats": {
            "total_pending": len(pending_review),
            "total_duplicates": len(duplicate_suspects),
            "pending_amount": total_pending_amount,
            "duplicate_amount": total_duplicate_amount
        }
    }

@api_router.get("/reconciliation/duplicates")
async def get_duplicate_pairs(
    user: dict = Depends(check_role([UserRole.ADMIN, UserRole.ACCOUNTANT]))
):
    """Get duplicate transaction pairs for review"""
    duplicates = await db.transactions.find(
        {"status": TransactionStatus.DUPLICATE_SUSPECT, "duplicate_of": {"$ne": None}},
        {"_id": 0}
    ).to_list(100)
    
    pairs = []
    for dup in duplicates:
        original = await db.transactions.find_one(
            {"id": dup.get("duplicate_of")},
            {"_id": 0}
        )
        if original:
            pairs.append({
                "duplicate": dup,
                "original": original,
                "confidence": dup.get("match_confidence", 0),
                "amount_match": dup.get("amount") == original.get("amount"),
                "date_diff_days": abs((datetime.strptime(dup.get("date", "2000-01-01"), "%Y-%m-%d") - 
                                      datetime.strptime(original.get("date", "2000-01-01"), "%Y-%m-%d")).days)
            })
    
    return {"pairs": pairs}

@api_router.put("/reconciliation/approve/{transaction_id}")
async def approve_transaction(
    transaction_id: str,
    category: Optional[str] = None,
    subcategory: Optional[str] = None,
    user: dict = Depends(check_role([UserRole.ADMIN, UserRole.ACCOUNTANT]))
):
    """Approve a transaction (mark as reconciled)"""
    update_data = {
        "status": TransactionStatus.APPROVED,
        "reviewed_by": user["id"],
        "reviewed_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Allow category correction
    if category:
        update_data["category"] = category
        update_data["is_deductible"] = SRI_CATEGORIES.get(category, {}).get("deductible", False)
    if subcategory:
        update_data["subcategory"] = subcategory
    
    result = await db.transactions.update_one(
        {"id": transaction_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Transacción no encontrada")
    
    return {"message": "Transacción aprobada", "status": TransactionStatus.APPROVED}

@api_router.put("/reconciliation/reject/{transaction_id}")
async def reject_transaction(
    transaction_id: str,
    reason: Optional[str] = None,
    user: dict = Depends(check_role([UserRole.ADMIN, UserRole.ACCOUNTANT]))
):
    """Reject a transaction (mark as invalid)"""
    result = await db.transactions.update_one(
        {"id": transaction_id},
        {"$set": {
            "status": TransactionStatus.REJECTED,
            "reviewed_by": user["id"],
            "reviewed_at": datetime.now(timezone.utc).isoformat(),
            "notes": reason or "Rechazado por contadora"
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Transacción no encontrada")
    
    return {"message": "Transacción rechazada", "status": TransactionStatus.REJECTED}

@api_router.put("/reconciliation/confirm-duplicate/{transaction_id}")
async def confirm_duplicate(
    transaction_id: str,
    keep_original: bool = True,
    user: dict = Depends(check_role([UserRole.ADMIN, UserRole.ACCOUNTANT]))
):
    """Confirm a transaction as duplicate"""
    # Get the duplicate transaction
    dup_tx = await db.transactions.find_one({"id": transaction_id}, {"_id": 0})
    if not dup_tx:
        raise HTTPException(status_code=404, detail="Transacción no encontrada")
    
    if keep_original:
        # Mark this one as confirmed duplicate (won't count in totals)
        await db.transactions.update_one(
            {"id": transaction_id},
            {"$set": {
                "status": TransactionStatus.DUPLICATE_CONFIRMED,
                "reviewed_by": user["id"],
                "reviewed_at": datetime.now(timezone.utc).isoformat(),
                "notes": "Confirmado como duplicado - no se cuenta en totales"
            }}
        )
        # Update original to have receipt/invoice
        if dup_tx.get("duplicate_of"):
            await db.transactions.update_one(
                {"id": dup_tx["duplicate_of"]},
                {"$set": {
                    "has_receipt": dup_tx.get("has_receipt", False) or True,
                    "has_invoice": dup_tx.get("has_invoice", False) or True
                }}
            )
    else:
        # Keep the duplicate, reject the original
        if dup_tx.get("duplicate_of"):
            await db.transactions.update_one(
                {"id": dup_tx["duplicate_of"]},
                {"$set": {
                    "status": TransactionStatus.DUPLICATE_CONFIRMED,
                    "reviewed_by": user["id"],
                    "reviewed_at": datetime.now(timezone.utc).isoformat()
                }}
            )
        # Approve this one
        await db.transactions.update_one(
            {"id": transaction_id},
            {"$set": {
                "status": TransactionStatus.APPROVED,
                "duplicate_of": None,
                "reviewed_by": user["id"],
                "reviewed_at": datetime.now(timezone.utc).isoformat()
            }}
        )
    
    return {"message": "Duplicado procesado", "kept_original": keep_original}

@api_router.put("/reconciliation/not-duplicate/{transaction_id}")
async def mark_not_duplicate(
    transaction_id: str,
    user: dict = Depends(check_role([UserRole.ADMIN, UserRole.ACCOUNTANT]))
):
    """Mark a transaction as NOT a duplicate (false positive)"""
    result = await db.transactions.update_one(
        {"id": transaction_id},
        {"$set": {
            "status": TransactionStatus.APPROVED,
            "duplicate_of": None,
            "match_confidence": None,
            "reviewed_by": user["id"],
            "reviewed_at": datetime.now(timezone.utc).isoformat(),
            "notes": "Revisado - no es duplicado"
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Transacción no encontrada")
    
    return {"message": "Marcado como no duplicado", "status": TransactionStatus.APPROVED}

@api_router.put("/reconciliation/bulk-approve")
async def bulk_approve_transactions(
    transaction_ids: List[str],
    user: dict = Depends(check_role([UserRole.ADMIN, UserRole.ACCOUNTANT]))
):
    """Bulk approve multiple transactions"""
    result = await db.transactions.update_many(
        {"id": {"$in": transaction_ids}},
        {"$set": {
            "status": TransactionStatus.APPROVED,
            "reviewed_by": user["id"],
            "reviewed_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": f"{result.modified_count} transacciones aprobadas"}

@api_router.get("/reconciliation/stats")
async def get_reconciliation_stats(
    user: dict = Depends(check_role([UserRole.ADMIN, UserRole.ACCOUNTANT]))
):
    """Get reconciliation statistics"""
    # Count by status
    pipeline = [
        {"$group": {"_id": "$status", "count": {"$sum": 1}, "total_amount": {"$sum": "$amount"}}}
    ]
    
    results = await db.transactions.aggregate(pipeline).to_list(10)
    
    stats = {
        "pending_review": 0,
        "approved": 0,
        "rejected": 0,
        "duplicate_suspect": 0,
        "duplicate_confirmed": 0,
        "total_pending_amount": 0,
        "total_approved_amount": 0
    }
    
    for r in results:
        status = r["_id"] or "pending_review"
        if status == TransactionStatus.PENDING_REVIEW:
            stats["pending_review"] = r["count"]
            stats["total_pending_amount"] = r["total_amount"]
        elif status == TransactionStatus.APPROVED:
            stats["approved"] = r["count"]
            stats["total_approved_amount"] = r["total_amount"]
        elif status == TransactionStatus.REJECTED:
            stats["rejected"] = r["count"]
        elif status == TransactionStatus.DUPLICATE_SUSPECT:
            stats["duplicate_suspect"] = r["count"]
        elif status == TransactionStatus.DUPLICATE_CONFIRMED:
            stats["duplicate_confirmed"] = r["count"]
    
    return stats

# ================= SPLIT TRANSACTIONS =================

@api_router.post("/transactions/split")
async def split_transaction(
    split_request: TransactionSplitRequest,
    user: dict = Depends(get_current_user)
):
    """Split a transaction into multiple categories (estilo QuickBooks)"""
    # Get original transaction
    original = await db.transactions.find_one(
        {"id": split_request.transaction_id, "user_id": user["id"]},
        {"_id": 0}
    )
    
    if not original:
        raise HTTPException(status_code=404, detail="Transacción no encontrada")
    
    # Validate split amounts
    total_split = sum(s.amount for s in split_request.splits)
    if abs(total_split - original["amount"]) > 0.01:
        raise HTTPException(
            status_code=400, 
            detail=f"La suma de los splits (${total_split:.2f}) debe ser igual al monto original (${original['amount']:.2f})"
        )
    
    # Create split transactions
    created_splits = []
    for i, split in enumerate(split_request.splits):
        split_id = str(uuid.uuid4())
        split_doc = {
            **original,
            "id": split_id,
            "amount": split.amount,
            "category": split.category,
            "subcategory": split.subcategory,
            "description": split.description or f"{original['description']} (Split {i+1})",
            "is_split": True,
            "parent_transaction_id": split_request.transaction_id,
            "is_deductible": SRI_CATEGORIES.get(split.category, {}).get("deductible", False),
            "status": TransactionStatus.PENDING_REVIEW,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.transactions.insert_one(split_doc)
        created_splits.append({k: v for k, v in split_doc.items() if k != "_id"})
    
    # Mark original as split (rejected/replaced)
    await db.transactions.update_one(
        {"id": split_request.transaction_id},
        {"$set": {
            "status": TransactionStatus.REJECTED,
            "notes": f"Dividido en {len(split_request.splits)} transacciones"
        }}
    )
    
    return {
        "message": f"Transacción dividida en {len(created_splits)} partes",
        "splits": created_splits
    }

# ================= ATTACHMENTS =================

@api_router.post("/transactions/{transaction_id}/attachments")
async def upload_attachment(
    transaction_id: str,
    file: UploadFile = File(...),
    attachment_type: str = Form("receipt"),  # receipt, invoice, other
    user: dict = Depends(get_current_user)
):
    """Upload attachment (receipt, invoice) to a transaction"""
    # Verify transaction exists
    transaction = await db.transactions.find_one(
        {"id": transaction_id, "user_id": user["id"]},
        {"_id": 0}
    )
    
    if not transaction:
        raise HTTPException(status_code=404, detail="Transacción no encontrada")
    
    # Save file
    file_ext = file.filename.split(".")[-1] if "." in file.filename else "jpg"
    filename = f"{transaction_id}_{attachment_type}_{uuid.uuid4().hex[:8]}.{file_ext}"
    file_path = UPLOADS_DIR / filename
    
    async with aiofiles.open(file_path, 'wb') as f:
        content = await file.read()
        await f.write(content)
    
    # Update transaction
    attachments = transaction.get("attachments", [])
    attachments.append(str(filename))
    
    update_data = {"attachments": attachments}
    if attachment_type == "receipt":
        update_data["has_receipt"] = True
    elif attachment_type == "invoice":
        update_data["has_invoice"] = True
    
    await db.transactions.update_one(
        {"id": transaction_id},
        {"$set": update_data}
    )
    
    return {
        "message": "Archivo adjuntado",
        "filename": filename,
        "type": attachment_type
    }

@api_router.get("/attachments/{filename}")
async def get_attachment(filename: str):
    """Download an attachment"""
    file_path = UPLOADS_DIR / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    
    return StreamingResponse(
        open(file_path, "rb"),
        media_type="application/octet-stream",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ================= EXPORT ENDPOINTS =================

@api_router.get("/export/transactions/excel")
async def export_transactions_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    """Export transactions to Excel"""
    query = {"user_id": user["id"], "status": {"$ne": TransactionStatus.DUPLICATE_CONFIRMED}}
    if start_date:
        query["date"] = {"$gte": start_date}
    if end_date:
        query.setdefault("date", {})["$lte"] = end_date
    
    transactions = await db.transactions.find(query, {"_id": 0}).sort("date", -1).to_list(10000)
    
    # Create Excel file
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)
    
    # Formats
    header_format = workbook.add_format({'bold': True, 'bg_color': '#2f9e44', 'font_color': 'white'})
    money_format = workbook.add_format({'num_format': '$#,##0.00'})
    date_format = workbook.add_format({'num_format': 'yyyy-mm-dd'})
    deductible_format = workbook.add_format({'bg_color': '#d3f9d8'})
    non_deductible_format = workbook.add_format({'bg_color': '#ffe3e3'})
    
    # Sheet 1: All transactions
    ws1 = workbook.add_worksheet("Transacciones")
    headers = ["Fecha", "Descripción", "Establecimiento", "Categoría", "Subcategoría", 
               "Monto", "Tipo", "Deducible SRI", "Estado", "Fuente"]
    
    for col, header in enumerate(headers):
        ws1.write(0, col, header, header_format)
    
    for row, t in enumerate(transactions, 1):
        ws1.write(row, 0, t.get("date", ""))
        ws1.write(row, 1, t.get("description", ""))
        ws1.write(row, 2, t.get("establishment", ""))
        ws1.write(row, 3, SRI_CATEGORIES.get(t.get("category", ""), {}).get("name", t.get("category", "")))
        ws1.write(row, 4, t.get("subcategory", ""))
        ws1.write(row, 5, t.get("amount", 0), money_format)
        ws1.write(row, 6, "Ingreso" if t.get("transaction_type") == "income" else "Gasto")
        ws1.write(row, 7, "Sí" if t.get("is_deductible") else "No")
        ws1.write(row, 8, STATUS_LABELS.get(t.get("status", ""), t.get("status", "")))
        ws1.write(row, 9, t.get("source_type", "manual"))
    
    ws1.autofilter(0, 0, len(transactions), len(headers) - 1)
    ws1.set_column(0, 0, 12)
    ws1.set_column(1, 1, 30)
    ws1.set_column(2, 4, 20)
    ws1.set_column(5, 5, 12)
    
    # Sheet 2: Summary by category
    ws2 = workbook.add_worksheet("Resumen SRI")
    ws2.write(0, 0, "Categoría", header_format)
    ws2.write(0, 1, "Total", header_format)
    ws2.write(0, 2, "Deducible", header_format)
    ws2.write(0, 3, "Límite SRI", header_format)
    ws2.write(0, 4, "% Usado", header_format)
    
    summary = {}
    for t in transactions:
        if t.get("transaction_type") == "expense":
            cat = t.get("category", "otros")
            summary[cat] = summary.get(cat, 0) + t.get("amount", 0)
    
    row = 1
    for cat, total in summary.items():
        cat_info = SRI_CATEGORIES.get(cat, {})
        limit = cat_info.get("limit_usd", 0)
        ws2.write(row, 0, cat_info.get("name", cat))
        ws2.write(row, 1, total, money_format)
        ws2.write(row, 2, "Sí" if cat_info.get("deductible") else "No")
        ws2.write(row, 3, limit, money_format)
        ws2.write(row, 4, f"{(total/limit*100):.1f}%" if limit > 0 else "N/A")
        row += 1
    
    workbook.close()
    output.seek(0)
    
    filename = f"transacciones_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# Labels for status
STATUS_LABELS = {
    TransactionStatus.PENDING_REVIEW: "Pendiente",
    TransactionStatus.APPROVED: "Aprobado",
    TransactionStatus.REJECTED: "Rechazado",
    TransactionStatus.DUPLICATE_SUSPECT: "Posible Duplicado",
    TransactionStatus.DUPLICATE_CONFIRMED: "Duplicado"
}

@api_router.get("/export/sri/pdf")
async def export_sri_pdf(
    year: Optional[int] = None,
    cargas_familiares: int = 3,
    user: dict = Depends(get_current_user)
):
    """Export SRI Gastos Personales report as PDF (Anexo)"""
    if not year:
        year = datetime.now().year
    
    start_date = f"{year}-01-01"
    end_date = f"{year}-12-31"
    
    # Get deductible transactions
    transactions = await db.transactions.find(
        {
            "user_id": user["id"],
            "date": {"$gte": start_date, "$lte": end_date},
            "transaction_type": "expense",
            "is_deductible": True,
            "status": {"$ne": TransactionStatus.DUPLICATE_CONFIRMED}
        },
        {"_id": 0}
    ).to_list(10000)
    
    # Calculate totals by category
    category_totals = {}
    for t in transactions:
        cat = t.get("category", "otros")
        category_totals[cat] = category_totals.get(cat, 0) + t.get("amount", 0)
    
    # Calculate limits
    cargas = min(cargas_familiares, 5)
    num_cbf = CARGAS_FAMILIARES_CBF.get(cargas, 7)
    limite_global = num_cbf * CANASTA_BASICA
    
    total_deductible = sum(category_totals.values())
    gastos_aplicables = min(total_deductible, limite_global)
    rebaja_ir = gastos_aplicables * PORCENTAJE_REBAJA_IR
    
    # Create PDF
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=16, spaceAfter=20)
    heading_style = ParagraphStyle('Heading', parent=styles['Heading2'], fontSize=12, spaceAfter=10)
    normal_style = styles['Normal']
    
    elements = []
    
    # Header
    elements.append(Paragraph("ANEXO DE GASTOS PERSONALES", title_style))
    elements.append(Paragraph(f"Año Fiscal: {year}", heading_style))
    elements.append(Spacer(1, 10))
    
    # Contribuyente info
    elements.append(Paragraph("DATOS DEL CONTRIBUYENTE", heading_style))
    info_data = [
        ["RUC:", CONTRIBUYENTE_INFO["ruc"]],
        ["Nombre:", CONTRIBUYENTE_INFO["nombre"]],
        ["Tipo:", CONTRIBUYENTE_INFO["tipo"]],
        ["Cargas Familiares:", str(cargas_familiares)],
    ]
    info_table = Table(info_data, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 20))
    
    # Deductible expenses table
    elements.append(Paragraph("GASTOS DEDUCIBLES POR CATEGORÍA", heading_style))
    
    table_data = [["Categoría", "Monto Gastado", "Límite", "% Usado"]]
    
    deductible_cats = ["alimentacion", "salud", "educacion", "vivienda", "vestimenta", "turismo"]
    for cat in deductible_cats:
        cat_info = SRI_CATEGORIES.get(cat, {})
        spent = category_totals.get(cat, 0)
        limit = cat_info.get("limit_usd", 0)
        pct = (spent / limit * 100) if limit > 0 else 0
        table_data.append([
            cat_info.get("name", cat),
            f"${spent:,.2f}",
            f"${limit:,.2f}",
            f"{pct:.1f}%"
        ])
    
    # Total row
    table_data.append(["TOTAL DEDUCIBLE", f"${total_deductible:,.2f}", f"${limite_global:,.2f}", 
                       f"{(total_deductible/limite_global*100):.1f}%" if limite_global > 0 else "0%"])
    
    expenses_table = Table(table_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1*inch])
    expenses_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2f9e44')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#d3f9d8')),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWHEIGHT', (0, 0), (-1, -1), 25),
    ]))
    elements.append(expenses_table)
    elements.append(Spacer(1, 20))
    
    # Rebaja calculation
    elements.append(Paragraph("CÁLCULO DE REBAJA", heading_style))
    rebaja_data = [
        ["Gastos Aplicables (menor entre total y límite):", f"${gastos_aplicables:,.2f}"],
        ["Porcentaje de Rebaja:", f"{PORCENTAJE_REBAJA_IR*100:.0f}%"],
        ["REBAJA ESTIMADA DE IMPUESTO A LA RENTA:", f"${rebaja_ir:,.2f}"],
    ]
    rebaja_table = Table(rebaja_data, colWidths=[4*inch, 2*inch])
    rebaja_table.setStyle(TableStyle([
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#2f9e44')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(rebaja_table)
    elements.append(Spacer(1, 30))
    
    # Footer
    elements.append(Paragraph(
        f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')} - FamilyFinance Ecuador",
        normal_style
    ))
    elements.append(Paragraph(
        "Este documento es un resumen para uso personal. Para la declaración oficial, use el formulario del SRI.",
        normal_style
    ))
    
    doc.build(elements)
    output.seek(0)
    
    filename = f"anexo_gastos_personales_{year}.pdf"
    
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ================= CATEGORIZATION RULES =================

@api_router.get("/categorization-rules")
async def get_categorization_rules(user: dict = Depends(get_current_user)):
    """Get all categorization rules"""
    # Get user custom rules
    custom_rules = await db.categorization_rules.find(
        {"user_id": user["id"]},
        {"_id": 0}
    ).to_list(100)
    
    return {
        "default_rules": DEFAULT_CATEGORIZATION_RULES,
        "custom_rules": custom_rules
    }

@api_router.post("/categorization-rules")
async def create_categorization_rule(
    rule: CategorizationRule,
    user: dict = Depends(get_current_user)
):
    """Create a custom categorization rule"""
    rule_id = str(uuid.uuid4())
    rule_doc = {
        "id": rule_id,
        "user_id": user["id"],
        **rule.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.categorization_rules.insert_one(rule_doc)
    
    return {"message": "Regla creada", "rule": {k: v for k, v in rule_doc.items() if k != "_id"}}

@api_router.delete("/categorization-rules/{rule_id}")
async def delete_categorization_rule(
    rule_id: str,
    user: dict = Depends(get_current_user)
):
    """Delete a custom categorization rule"""
    result = await db.categorization_rules.delete_one({"id": rule_id, "user_id": user["id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Regla no encontrada")
    return {"message": "Regla eliminada"}

@api_router.post("/transactions/auto-categorize")
async def auto_categorize_transaction(
    description: str,
    establishment: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    """Test auto-categorization for a description"""
    # First check custom rules
    custom_rules = await db.categorization_rules.find(
        {"user_id": user["id"], "is_active": True},
        {"_id": 0}
    ).to_list(100)
    
    text = f"{description} {establishment or ''}".lower()
    
    for rule in custom_rules:
        for keyword in rule.get("keywords", []):
            if keyword.lower() in text:
                return {
                    "category": rule["category"],
                    "subcategory": rule["subcategory"],
                    "auto_categorized": True,
                    "matched_keyword": keyword,
                    "rule_type": "custom"
                }
    
    # Then check default rules
    result = apply_categorization_rules(description, establishment or "")
    if result["auto_categorized"]:
        result["rule_type"] = "default"
        return result
    
    return {
        "category": None,
        "subcategory": None,
        "auto_categorized": False,
        "message": "No se encontró regla de categorización"
    }

# ================= USERS MANAGEMENT (ADMIN ONLY) =================

@api_router.get("/users", response_model=List[UserResponse])
async def get_users(user: dict = Depends(check_role([UserRole.ADMIN]))):
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(100)
    return [UserResponse(**u) for u in users]

@api_router.put("/users/{user_id}/role")
async def update_user_role(
    user_id: str,
    role: str,
    user: dict = Depends(check_role([UserRole.ADMIN]))
):
    if role not in [UserRole.ADMIN, UserRole.ACCOUNTANT, UserRole.SPOUSE]:
        raise HTTPException(status_code=400, detail="Rol inválido")
    
    result = await db.users.update_one({"id": user_id}, {"$set": {"role": role}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    return {"message": "Rol actualizado"}

# ================= INCOME MANAGEMENT (Manual Entry) =================

@api_router.get("/income")
async def get_incomes(
    year: Optional[int] = None,
    distribution: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    """Get all income entries"""
    query = {"user_id": user["id"]}
    if year:
        query["date"] = {"$regex": f"^{year}"}
    if distribution:
        query["distribution"] = distribution
    
    incomes = await db.incomes.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    return incomes

@api_router.post("/income")
async def create_income(
    income: IncomeEntry,
    user: dict = Depends(get_current_user)
):
    """Create a new income entry with distribution"""
    if user["role"] not in [UserRole.ADMIN, UserRole.SPOUSE]:
        raise HTTPException(status_code=403, detail="No autorizado para crear ingresos")
    
    income_doc = {
        "id": str(uuid.uuid4()),
        "user_id": user["id"],
        "amount": income.amount,
        "date": income.date,
        "distribution": income.distribution,
        "concept": income.concept,
        "description": income.description,
        "is_recurring": income.is_recurring,
        "payment_method": income.payment_method,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.incomes.insert_one(income_doc)
    del income_doc["_id"]
    
    return income_doc

@api_router.put("/income/{income_id}")
async def update_income(
    income_id: str,
    income: IncomeEntry,
    user: dict = Depends(get_current_user)
):
    """Update an income entry"""
    if user["role"] not in [UserRole.ADMIN, UserRole.SPOUSE]:
        raise HTTPException(status_code=403, detail="No autorizado")
    
    update_data = {
        "amount": income.amount,
        "date": income.date,
        "distribution": income.distribution,
        "concept": income.concept,
        "description": income.description,
        "is_recurring": income.is_recurring,
        "payment_method": income.payment_method
    }
    
    result = await db.incomes.update_one({"id": income_id, "user_id": user["id"]}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Ingreso no encontrado")
    
    return {"message": "Ingreso actualizado"}

@api_router.delete("/income/{income_id}")
async def delete_income(
    income_id: str,
    user: dict = Depends(get_current_user)
):
    """Delete an income entry"""
    if user["role"] not in [UserRole.ADMIN, UserRole.SPOUSE]:
        raise HTTPException(status_code=403, detail="No autorizado")
    
    result = await db.incomes.delete_one({"id": income_id, "user_id": user["id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Ingreso no encontrado")
    
    return {"message": "Ingreso eliminado"}

@api_router.get("/income/summary")
async def get_income_summary(
    year: Optional[int] = None,
    user: dict = Depends(get_current_user)
):
    """Get income summary by distribution"""
    current_year = year or datetime.now().year
    query = {"user_id": user["id"], "date": {"$regex": f"^{current_year}"}}
    
    incomes = await db.incomes.find(query, {"_id": 0}).to_list(1000)
    
    by_distribution = {}
    by_concept = {}
    total = 0
    
    for income in incomes:
        dist = income.get("distribution", "Personal")
        concept = income.get("concept", "Otros")
        amount = income.get("amount", 0)
        
        by_distribution[dist] = by_distribution.get(dist, 0) + amount
        by_concept[concept] = by_concept.get(concept, 0) + amount
        total += amount
    
    return {
        "total": total,
        "by_distribution": by_distribution,
        "by_concept": by_concept,
        "year": current_year,
        "count": len(incomes)
    }

# ================= PERSONAL BUDGET (User's Custom Categories) =================

@api_router.get("/budget/categories")
async def get_budget_categories_endpoint(user: dict = Depends(get_current_user)):
    """Get personal budget categories (from Excel structure)"""
    budget_cats = get_budget_categories(user)
    budget_goals_data = get_budget_goals(user)
    return {
        "categories": budget_cats,
        "payment_methods": PAYMENT_METHODS,
        "goals": budget_goals_data,
        "income_sources": INCOME_SOURCES,
        "income_concepts": INCOME_CONCEPTS
    }

@api_router.get("/budget/personal")
async def get_personal_budget(
    year: Optional[int] = None,
    month: Optional[int] = None,
    user: dict = Depends(get_current_user)
):
    """Get personal budget with actuals vs planned"""
    current_year = year or datetime.now().year
    
    # Get planned budget from database
    budget_query = {"user_id": user["id"], "year": current_year}
    planned_budget = await db.personal_budgets.find_one(budget_query, {"_id": 0})
    
    # Get actual expenses by budget category
    expense_query = {"user_id": user["id"], "date": {"$regex": f"^{current_year}"}}
    if month:
        expense_query["date"] = {"$regex": f"^{current_year}-{str(month).zfill(2)}"}
    
    transactions = await db.transactions.find(
        expense_query, 
        {"_id": 0, "amount": 1, "budget_category": 1, "category": 1, "date": 1}
    ).to_list(10000)
    
    # Get incomes
    incomes = await db.incomes.find(expense_query, {"_id": 0}).to_list(1000)
    total_income = sum(i.get("amount", 0) for i in incomes)
    
    # Calculate actuals by budget category
    actuals = {}
    for t in transactions:
        cat = t.get("budget_category") or t.get("category") or "otros"
        actuals[cat] = actuals.get(cat, 0) + t.get("amount", 0)
    
    total_expenses = sum(actuals.values())
    
    # Calculate goal progress
    budget_goals_data = get_budget_goals(user)
    goal_progress = {
        "gastos_fijos": {
            "target_percent": budget_goals_data["gastos_fijos_target"],
            "actual_percent": total_expenses / total_income if total_income > 0 else 0,
            "status": "on_track" if total_income > 0 and (total_expenses / total_income) <= budget_goals_data["gastos_fijos_target"]["max"] else "over"
        },
        "gastos_libres": {
            "target_annual": budget_goals_data["gastos_libres_max_annual"],
            "actual_annual": actuals.get("gastos_libres", 0),
            "remaining": budget_goals_data["gastos_libres_max_annual"] - actuals.get("gastos_libres", 0)
        }
    }
    
    budget_cats = get_budget_categories(user)
    return {
        "year": current_year,
        "month": month,
        "total_income": total_income,
        "total_expenses": total_expenses,
        "balance": total_income - total_expenses,
        "by_category": actuals,
        "planned": planned_budget.get("categories", {}) if planned_budget else {},
        "goal_progress": goal_progress,
        "categories_config": budget_cats
    }

@api_router.post("/budget/personal")
async def save_personal_budget(
    budget_data: dict,
    user: dict = Depends(get_current_user)
):
    """Save personal budget configuration"""
    if user["role"] not in [UserRole.ADMIN, UserRole.SPOUSE, "demo"]:
        raise HTTPException(status_code=403, detail="No autorizado")
    
    year = budget_data.get("year", datetime.now().year)
    income_struct = get_income_structure(user)
    budget_summ = get_budget_summary(user)
    budget_goals_data = get_budget_goals(user)
    
    budget_doc = {
        "user_id": user["id"],
        "year": year,
        "categories": budget_data.get("categories", {}),
        "income_projection": budget_data.get("income_projection", income_struct),
        "savings_goal": budget_data.get("savings_goal", budget_summ["ahorro_esperado"]),
        "investment_goal": budget_data.get("investment_goal", budget_summ["inversion_esperada"]),
        "goals": budget_data.get("goals", budget_goals_data),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.personal_budgets.update_one(
        {"user_id": user["id"], "year": year},
        {"$set": budget_doc},
        upsert=True
    )
    
    return {"message": "Presupuesto guardado", "year": year}

@api_router.get("/budget/config")
async def get_budget_config(
    year: Optional[int] = None,
    user: dict = Depends(get_current_user)
):
    """Get user's editable budget configuration"""
    current_year = year or datetime.now().year
    
    # Try to get saved budget
    saved = await db.personal_budgets.find_one(
        {"user_id": user["id"], "year": current_year},
        {"_id": 0}
    )
    
    # Get appropriate budget categories for user type
    base_budget_cats = get_budget_categories(user)
    
    # Merge saved categories with defaults to preserve subcategories
    merged_categories = dict(base_budget_cats)  # Start with defaults
    
    if saved and saved.get("categories"):
        saved_cats = saved["categories"]
        for key, default_cat in base_budget_cats.items():
            if key in saved_cats:
                # User has this category, merge with defaults
                merged_cat = dict(default_cat)  # Start with default
                saved_cat = saved_cats[key]
                
                # Override with saved values
                merged_cat["monthly_budget"] = saved_cat.get("monthly_budget", default_cat.get("monthly_budget", 0))
                merged_cat["annual_budget"] = merged_cat["monthly_budget"] * 12
                merged_cat["name"] = saved_cat.get("name", default_cat.get("name", key))
                merged_cat["type"] = saved_cat.get("type", default_cat.get("type", "variable"))
                merged_cat["is_recurring"] = saved_cat.get("is_recurring", default_cat.get("is_recurring", False))
                
                # Merge subcategories - keep defaults but override values if user set them
                default_subs = default_cat.get("subcategories", {})
                saved_subs = saved_cat.get("subcategories", {})
                if isinstance(default_subs, dict):
                    merged_subs = dict(default_subs)
                    if isinstance(saved_subs, dict):
                        merged_subs.update(saved_subs)
                    merged_cat["subcategories"] = merged_subs
                else:
                    merged_cat["subcategories"] = saved_subs if saved_subs else {}
                    
                merged_categories[key] = merged_cat
            else:
                # User doesn't have this category, keep default
                merged_categories[key] = default_cat
        
        # Add any custom categories the user created
        for key, cat in saved_cats.items():
            if key not in base_budget_cats:
                merged_categories[key] = cat
    
    return {
        "year": current_year,
        "categories": merged_categories,
        "income_projection": saved.get("income_projection", get_income_structure(user)) if saved else get_income_structure(user),
        "savings_goal": saved.get("savings_goal", get_budget_summary(user)["ahorro_esperado"]) if saved else get_budget_summary(user)["ahorro_esperado"],
        "investment_goal": saved.get("investment_goal", get_budget_summary(user)["inversion_esperada"]) if saved else get_budget_summary(user)["inversion_esperada"],
        "is_custom": saved is not None
    }

@api_router.put("/budget/category/{category_key}")
async def update_budget_category(
    category_key: str,
    category_data: dict,
    user: dict = Depends(get_current_user)
):
    """Update a single budget category"""
    if user["role"] not in [UserRole.ADMIN, UserRole.SPOUSE]:
        raise HTTPException(status_code=403, detail="No autorizado")
    
    year = datetime.now().year
    
    # Get or create budget
    budget = await db.personal_budgets.find_one(
        {"user_id": user["id"], "year": year}
    )
    
    if not budget:
        budget = {
            "user_id": user["id"],
            "year": year,
            "categories": dict(BUDGET_CATEGORIES),
            "income_projection": dict(INCOME_STRUCTURE),
            "savings_goal": dict(BUDGET_SUMMARY["ahorro_esperado"]),
            "investment_goal": dict(BUDGET_SUMMARY["inversion_esperada"])
        }
    
    # Update category
    categories = budget.get("categories", {})
    if category_key in categories:
        categories[category_key].update(category_data)
    else:
        categories[category_key] = category_data
    
    await db.personal_budgets.update_one(
        {"user_id": user["id"], "year": year},
        {"$set": {"categories": categories, "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    
    return {"message": f"Categoría {category_key} actualizada"}

@api_router.post("/budget/category")
async def add_budget_category(
    category_data: dict,
    user: dict = Depends(get_current_user)
):
    """Add a new budget category"""
    if user["role"] not in [UserRole.ADMIN, UserRole.SPOUSE]:
        raise HTTPException(status_code=403, detail="No autorizado")
    
    category_key = category_data.get("key")
    if not category_key:
        raise HTTPException(status_code=400, detail="Se requiere key de categoría")
    
    year = datetime.now().year
    
    await db.personal_budgets.update_one(
        {"user_id": user["id"], "year": year},
        {"$set": {f"categories.{category_key}": category_data, "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    
    return {"message": f"Categoría {category_key} añadida"}

@api_router.delete("/budget/category/{category_key}")
async def delete_budget_category(
    category_key: str,
    user: dict = Depends(get_current_user)
):
    """Delete a budget category"""
    if user["role"] not in [UserRole.ADMIN, UserRole.SPOUSE]:
        raise HTTPException(status_code=403, detail="No autorizado")
    
    year = datetime.now().year
    
    await db.personal_budgets.update_one(
        {"user_id": user["id"], "year": year},
        {"$unset": {f"categories.{category_key}": ""}}
    )
    
    return {"message": f"Categoría {category_key} eliminada"}

# ================= DEFERRED PAYMENTS (DIFERIDOS) =================

class DeferredPayment(BaseModel):
    description: str
    total_amount: float
    monthly_payment: float
    remaining_installments: int
    total_installments: int
    card_id: Optional[str] = None
    card_name: Optional[str] = None
    start_date: Optional[str] = None

@api_router.get("/deferred-payments")
async def get_deferred_payments(user: dict = Depends(get_current_user)):
    """Get all deferred/installment payments"""
    payments = await db.deferred_payments.find(
        {"user_id": user["id"], "remaining_installments": {"$gt": 0}},
        {"_id": 0}
    ).sort("card_name", 1).to_list(100)
    
    total_remaining = sum(p.get("monthly_payment", 0) * p.get("remaining_installments", 0) for p in payments)
    total_monthly = sum(p.get("monthly_payment", 0) for p in payments)
    
    return {
        "payments": payments,
        "total_remaining": total_remaining,
        "total_monthly_obligation": total_monthly,
        "count": len(payments)
    }

@api_router.post("/deferred-payments")
async def create_deferred_payment(
    payment: DeferredPayment,
    user: dict = Depends(get_current_user)
):
    """Create a new deferred payment"""
    payment_doc = {
        "id": str(uuid.uuid4()),
        "user_id": user["id"],
        **payment.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.deferred_payments.insert_one(payment_doc)
    
    return {"message": "Diferido creado", "payment": {k: v for k, v in payment_doc.items() if k != "_id"}}

@api_router.put("/deferred-payments/{payment_id}")
async def update_deferred_payment(
    payment_id: str,
    update_data: dict,
    user: dict = Depends(get_current_user)
):
    """Update a deferred payment (e.g., reduce remaining installments)"""
    result = await db.deferred_payments.update_one(
        {"id": payment_id, "user_id": user["id"]},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Diferido no encontrado")
    
    return {"message": "Diferido actualizado"}

@api_router.delete("/deferred-payments/{payment_id}")
async def delete_deferred_payment(
    payment_id: str,
    user: dict = Depends(get_current_user)
):
    """Delete a deferred payment"""
    result = await db.deferred_payments.delete_one({"id": payment_id, "user_id": user["id"]})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Diferido no encontrado")
    
    return {"message": "Diferido eliminado"}

# ================= TRANSACTIONS GROUPED BY ESTABLISHMENT =================

@api_router.get("/transactions/grouped")
async def get_transactions_grouped(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    """Get transactions grouped by establishment and date (accordion view)"""
    query = {"user_id": user["id"], "transaction_type": "expense"}
    if start_date:
        query["date"] = {"$gte": start_date}
    if end_date:
        query.setdefault("date", {})["$lte"] = end_date
    
    transactions = await db.transactions.find(query, {"_id": 0}).sort("date", -1).to_list(10000)
    
    # Group by establishment + date
    groups = {}
    for t in transactions:
        # Create group key: establishment + date
        establishment = t.get("establishment") or t.get("description", "Sin establecimiento")
        date = t.get("date", "")
        group_key = f"{establishment}|{date}"
        
        if group_key not in groups:
            groups[group_key] = {
                "establishment": establishment,
                "date": date,
                "total": 0,
                "items": [],
                "category": t.get("category"),
                "payment_method": t.get("payment_method"),
                "attachments": []
            }
        
        groups[group_key]["total"] += t.get("amount", 0)
        groups[group_key]["items"].append({
            "id": t.get("id"),
            "description": t.get("description"),
            "amount": t.get("amount"),
            "category": t.get("category"),
            "subcategory": t.get("subcategory")
        })
        
        if t.get("attachments"):
            groups[group_key]["attachments"].extend(t["attachments"])
    
    # Convert to list and sort by date
    result = list(groups.values())
    result.sort(key=lambda x: x["date"], reverse=True)
    
    return result

# ================= CREDIT CARDS & DEBT MANAGEMENT =================

@api_router.get("/credit-cards")
async def get_credit_cards(user: dict = Depends(get_current_user)):
    """Get all credit cards for user"""
    cards = await db.credit_cards.find({"user_id": user["id"]}, {"_id": 0}).to_list(100)
    
    # Calculate available credit for each card
    for card in cards:
        card["available_credit"] = card.get("credit_limit", 0) - card.get("current_balance", 0)
    
    return cards

@api_router.post("/credit-cards")
async def create_credit_card(card: CreditCard, user: dict = Depends(get_current_user)):
    """Add a new credit card"""
    if user["role"] not in [UserRole.ADMIN, UserRole.SPOUSE]:
        raise HTTPException(status_code=403, detail="No autorizado")
    
    card_doc = {
        "id": str(uuid.uuid4()),
        "user_id": user["id"],
        **card.model_dump(),
        "available_credit": card.credit_limit - card.current_balance,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": None
    }
    
    await db.credit_cards.insert_one(card_doc)
    del card_doc["_id"]
    
    return card_doc

@api_router.put("/credit-cards/{card_id}")
async def update_credit_card(card_id: str, card: CreditCard, user: dict = Depends(get_current_user)):
    """Update credit card details"""
    if user["role"] not in [UserRole.ADMIN, UserRole.SPOUSE]:
        raise HTTPException(status_code=403, detail="No autorizado")
    
    update_data = {
        **card.model_dump(),
        "available_credit": card.credit_limit - card.current_balance,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    result = await db.credit_cards.update_one(
        {"id": card_id, "user_id": user["id"]},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Tarjeta no encontrada")
    
    return {"message": "Tarjeta actualizada"}

@api_router.delete("/credit-cards/{card_id}")
async def delete_credit_card(card_id: str, user: dict = Depends(get_current_user)):
    """Delete a credit card"""
    if user["role"] not in [UserRole.ADMIN, UserRole.SPOUSE]:
        raise HTTPException(status_code=403, detail="No autorizado")
    
    result = await db.credit_cards.delete_one({"id": card_id, "user_id": user["id"]})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Tarjeta no encontrada")
    
    return {"message": "Tarjeta eliminada"}

@api_router.get("/debt/summary")
async def get_debt_summary(user: dict = Depends(get_current_user)):
    """Get debt summary across all cards"""
    cards = await db.credit_cards.find({"user_id": user["id"]}, {"_id": 0}).to_list(100)
    
    total_debt = sum(c.get("current_balance", 0) for c in cards)
    total_limit = sum(c.get("credit_limit", 0) for c in cards)
    total_minimum = sum(c.get("minimum_payment", 0) for c in cards)
    
    # Calculate weighted average APR
    if total_debt > 0:
        weighted_apr = sum(c.get("current_balance", 0) * c.get("apr", 0) for c in cards) / total_debt
    else:
        weighted_apr = 0
    
    # Sort by APR for avalanche strategy
    cards_by_apr = sorted(cards, key=lambda x: x.get("apr", 0), reverse=True)
    
    return {
        "total_debt": total_debt,
        "total_credit_limit": total_limit,
        "total_available_credit": total_limit - total_debt,
        "total_minimum_payment": total_minimum,
        "weighted_average_apr": round(weighted_apr, 2),
        "utilization_rate": round((total_debt / total_limit * 100) if total_limit > 0 else 0, 1),
        "cards_count": len(cards),
        "cards": cards,
        "highest_apr_card": cards_by_apr[0] if cards_by_apr else None
    }

@api_router.post("/debt/snowball-plan")
async def calculate_snowball_plan(
    plan: SnowballPlan,
    user: dict = Depends(get_current_user)
):
    """Calculate debt payoff plan using Avalanche or Snowball method"""
    cards = await db.credit_cards.find(
        {"user_id": user["id"], "current_balance": {"$gt": 0}},
        {"_id": 0}
    ).to_list(100)
    
    if not cards:
        return {"message": "No hay deudas activas", "plan": [], "months_to_payoff": 0}
    
    # Sort cards based on strategy
    if plan.strategy == "avalanche":
        # Pay highest interest first (mathematically optimal)
        cards = sorted(cards, key=lambda x: x.get("apr", 0), reverse=True)
    else:
        # Pay smallest balance first (psychologically motivating)
        cards = sorted(cards, key=lambda x: x.get("current_balance", 0))
    
    # Calculate payoff plan
    payoff_plan = []
    total_interest_saved = 0
    months = 0
    max_months = 120  # 10 years max
    
    # Clone cards for simulation
    import copy
    sim_cards = copy.deepcopy(cards)
    
    while any(c["current_balance"] > 0 for c in sim_cards) and months < max_months:
        months += 1
        month_plan = {"month": months, "payments": [], "total_payment": 0}
        
        # First, pay minimum on all cards
        for card in sim_cards:
            if card["current_balance"] > 0:
                # Add monthly interest
                monthly_interest = card["current_balance"] * (card["apr"] / 100 / 12)
                card["current_balance"] += monthly_interest
                
                # Pay minimum
                min_payment = min(card["minimum_payment"], card["current_balance"])
                card["current_balance"] -= min_payment
                month_plan["payments"].append({
                    "card": card["name"],
                    "payment": min_payment,
                    "remaining": card["current_balance"],
                    "type": "minimum"
                })
                month_plan["total_payment"] += min_payment
        
        # Apply extra payment to target card (first with balance > 0)
        extra_remaining = plan.extra_payment
        for card in sim_cards:
            if card["current_balance"] > 0 and extra_remaining > 0:
                extra_to_pay = min(extra_remaining, card["current_balance"])
                card["current_balance"] -= extra_to_pay
                extra_remaining -= extra_to_pay
                
                # Update the payment in month_plan
                for p in month_plan["payments"]:
                    if p["card"] == card["name"]:
                        p["payment"] += extra_to_pay
                        p["remaining"] = card["current_balance"]
                        p["type"] = "extra" if extra_to_pay > 0 else p["type"]
                        break
                
                month_plan["total_payment"] += extra_to_pay
                
                if card["current_balance"] <= 0:
                    card["current_balance"] = 0
                    # Card paid off! Continue to next card
        
        payoff_plan.append(month_plan)
    
    # Calculate total paid and interest
    total_paid = sum(m["total_payment"] for m in payoff_plan)
    original_debt = sum(c.get("current_balance", 0) for c in cards)
    total_interest = total_paid - original_debt
    
    return {
        "strategy": plan.strategy,
        "strategy_name": "Avalanche (mayor interés primero)" if plan.strategy == "avalanche" else "Snowball (menor saldo primero)",
        "months_to_payoff": months,
        "years_to_payoff": round(months / 12, 1),
        "total_paid": round(total_paid, 2),
        "total_interest": round(total_interest, 2),
        "original_debt": round(original_debt, 2),
        "extra_payment_monthly": plan.extra_payment,
        "payoff_order": [c["name"] for c in cards],
        "monthly_plan": payoff_plan[:12],  # First 12 months detail
        "recommendation": f"Paga primero {cards[0]['name']} ({cards[0]['apr']}% APR) para ahorrar en intereses" if cards else None
    }

@api_router.post("/debt/payment")
async def record_debt_payment(payment: DebtPayment, user: dict = Depends(get_current_user)):
    """Record a payment to a credit card"""
    if user["role"] not in [UserRole.ADMIN, UserRole.SPOUSE]:
        raise HTTPException(status_code=403, detail="No autorizado")
    
    # Get card
    card = await db.credit_cards.find_one({"id": payment.card_id, "user_id": user["id"]})
    if not card:
        raise HTTPException(status_code=404, detail="Tarjeta no encontrada")
    
    # Update balance
    new_balance = max(0, card["current_balance"] - payment.amount)
    
    await db.credit_cards.update_one(
        {"id": payment.card_id},
        {"$set": {
            "current_balance": new_balance,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Record payment history
    payment_doc = {
        "id": str(uuid.uuid4()),
        "user_id": user["id"],
        "card_id": payment.card_id,
        "card_name": card["name"],
        "amount": payment.amount,
        "payment_type": payment.payment_type,
        "date": payment.date,
        "balance_after": new_balance,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.debt_payments.insert_one(payment_doc)
    
    return {"message": "Pago registrado", "new_balance": new_balance}

# ================= CASH FLOW PLANNING =================

@api_router.get("/scheduled-payments")
async def get_scheduled_payments(user: dict = Depends(get_current_user)):
    """Get all scheduled payments"""
    payments = await db.scheduled_payments.find({"user_id": user["id"]}, {"_id": 0}).to_list(100)
    
    # Calculate next due date for each payment
    today = datetime.now(timezone.utc)
    current_month = today.month
    current_year = today.year
    
    for payment in payments:
        due_day = payment.get("due_day", 1)
        # If due day has passed this month, next is next month
        if today.day > due_day:
            next_month = current_month + 1 if current_month < 12 else 1
            next_year = current_year if current_month < 12 else current_year + 1
        else:
            next_month = current_month
            next_year = current_year
        
        payment["next_due_date"] = f"{next_year}-{str(next_month).zfill(2)}-{str(due_day).zfill(2)}"
        
        # Check if payment is due soon (within reminder days)
        days_until_due = (datetime(next_year, next_month, min(due_day, 28)) - today.replace(tzinfo=None)).days
        payment["is_due_soon"] = days_until_due <= payment.get("reminder_days_before", 2)
        payment["days_until_due"] = days_until_due
    
    # Sort by next due date
    payments.sort(key=lambda x: x.get("next_due_date", ""))
    
    return payments

@api_router.post("/scheduled-payments")
async def create_scheduled_payment(payment: ScheduledPayment, user: dict = Depends(get_current_user)):
    """Create a new scheduled payment"""
    if user["role"] not in [UserRole.ADMIN, UserRole.SPOUSE]:
        raise HTTPException(status_code=403, detail="No autorizado")
    
    payment_doc = {
        "id": str(uuid.uuid4()),
        "user_id": user["id"],
        **payment.model_dump(),
        "last_paid_date": None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.scheduled_payments.insert_one(payment_doc)
    del payment_doc["_id"]
    
    return payment_doc

@api_router.put("/scheduled-payments/{payment_id}")
async def update_scheduled_payment(
    payment_id: str, 
    payment: ScheduledPayment, 
    user: dict = Depends(get_current_user)
):
    """Update a scheduled payment"""
    if user["role"] not in [UserRole.ADMIN, UserRole.SPOUSE]:
        raise HTTPException(status_code=403, detail="No autorizado")
    
    result = await db.scheduled_payments.update_one(
        {"id": payment_id, "user_id": user["id"]},
        {"$set": payment.model_dump()}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Pago no encontrado")
    
    return {"message": "Pago actualizado"}

@api_router.delete("/scheduled-payments/{payment_id}")
async def delete_scheduled_payment(payment_id: str, user: dict = Depends(get_current_user)):
    """Delete a scheduled payment"""
    if user["role"] not in [UserRole.ADMIN, UserRole.SPOUSE]:
        raise HTTPException(status_code=403, detail="No autorizado")
    
    result = await db.scheduled_payments.delete_one({"id": payment_id, "user_id": user["id"]})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Pago no encontrado")
    
    return {"message": "Pago eliminado"}

@api_router.post("/scheduled-payments/{payment_id}/mark-paid")
async def mark_payment_paid(payment_id: str, user: dict = Depends(get_current_user)):
    """Mark a scheduled payment as paid this month"""
    if user["role"] not in [UserRole.ADMIN, UserRole.SPOUSE]:
        raise HTTPException(status_code=403, detail="No autorizado")
    
    result = await db.scheduled_payments.update_one(
        {"id": payment_id, "user_id": user["id"]},
        {"$set": {"last_paid_date": datetime.now(timezone.utc).strftime("%Y-%m-%d")}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Pago no encontrado")
    
    return {"message": "Marcado como pagado"}

# ================= EXPECTED INCOME ENDPOINTS =================

@api_router.get("/expected-income")
async def get_expected_income(user: dict = Depends(get_current_user)):
    """Get all expected income for the user"""
    items = await db.expected_income.find(
        {"user_id": user["id"]}
    ).sort("expected_date", 1).to_list(100)
    
    # Update overdue status
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    for item in items:
        if item.get("status") == "pending" and item.get("expected_date", "") < today:
            item["status"] = "overdue"
    
    total_pending = sum(i.get("amount", 0) for i in items if i.get("status") in ["pending", "overdue"])
    
    return {
        "items": [{k: v for k, v in item.items() if k != "_id"} for item in items],
        "total_pending": total_pending,
        "count": len(items)
    }

@api_router.post("/expected-income")
async def create_expected_income(
    income: ExpectedIncomeCreate,
    user: dict = Depends(get_current_user)
):
    """Create a new expected income entry"""
    doc = {
        "id": str(uuid.uuid4()),
        "user_id": user["id"],
        **income.model_dump(),
        "status": "pending",
        "linked_transaction_id": None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.expected_income.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}

@api_router.put("/expected-income/{income_id}")
async def update_expected_income(
    income_id: str,
    updates: dict,
    user: dict = Depends(get_current_user)
):
    """Update an expected income entry"""
    result = await db.expected_income.update_one(
        {"id": income_id, "user_id": user["id"]},
        {"$set": updates}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Ingreso no encontrado")
    return {"message": "Actualizado"}

@api_router.put("/expected-income/{income_id}/mark-received")
async def mark_income_received(
    income_id: str,
    user: dict = Depends(get_current_user)
):
    """Mark expected income as received and create actual income transaction"""
    income = await db.expected_income.find_one({"id": income_id, "user_id": user["id"]})
    if not income:
        raise HTTPException(status_code=404, detail="Ingreso no encontrado")
    
    # Create actual income transaction
    transaction_id = str(uuid.uuid4())
    transaction_doc = {
        "id": transaction_id,
        "user_id": user["id"],
        "amount": income["amount"],
        "description": income["description"],
        "category": "ingreso",
        "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "transaction_type": "income",
        "source": income.get("source", "personal"),
        "status": "approved",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.transactions.insert_one(transaction_doc)
    
    # Update expected income status
    await db.expected_income.update_one(
        {"id": income_id},
        {"$set": {
            "status": "received",
            "linked_transaction_id": transaction_id,
            "received_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "Ingreso recibido", "transaction_id": transaction_id}

@api_router.delete("/expected-income/{income_id}")
async def delete_expected_income(
    income_id: str,
    user: dict = Depends(get_current_user)
):
    result = await db.expected_income.delete_one({"id": income_id, "user_id": user["id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Ingreso no encontrado")
    return {"message": "Eliminado"}

# ================= ACCOUNTS RECEIVABLE ENDPOINTS =================

@api_router.get("/accounts-receivable")
async def get_accounts_receivable(user: dict = Depends(get_current_user)):
    """Get all accounts receivable"""
    items = await db.accounts_receivable.find(
        {"user_id": user["id"]}
    ).sort("due_date", 1).to_list(100)
    
    # Update status based on due date
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    for item in items:
        if item.get("status") == "pending" and item.get("due_date", "") < today:
            item["status"] = "overdue"
    
    total_pending = sum(i.get("amount", 0) - i.get("amount_paid", 0) for i in items if i.get("status") in ["pending", "overdue", "partial"])
    
    return {
        "items": [{k: v for k, v in item.items() if k != "_id"} for item in items],
        "total_pending": total_pending,
        "count": len(items)
    }

@api_router.post("/accounts-receivable")
async def create_account_receivable(
    account: AccountReceivableCreate,
    user: dict = Depends(get_current_user)
):
    """Create a new account receivable"""
    doc = {
        "id": str(uuid.uuid4()),
        "user_id": user["id"],
        **account.model_dump(),
        "amount_paid": 0,
        "status": "pending",
        "payment_history": [],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.accounts_receivable.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}

@api_router.put("/accounts-receivable/{account_id}/payment")
async def record_receivable_payment(
    account_id: str,
    payment_data: dict,
    user: dict = Depends(get_current_user)
):
    """Record a payment for an account receivable"""
    amount = payment_data.get("amount", 0)
    if amount <= 0:
        raise HTTPException(status_code=400, detail="Monto debe ser mayor a 0")
    
    account = await db.accounts_receivable.find_one({"id": account_id, "user_id": user["id"]})
    if not account:
        raise HTTPException(status_code=404, detail="Cuenta no encontrada")
    
    new_amount_paid = account.get("amount_paid", 0) + amount
    new_status = "paid" if new_amount_paid >= account["amount"] else "partial"
    
    payment_record = {
        "amount": amount,
        "date": datetime.now(timezone.utc).isoformat(),
        "id": str(uuid.uuid4())
    }
    
    await db.accounts_receivable.update_one(
        {"id": account_id},
        {
            "$set": {"amount_paid": new_amount_paid, "status": new_status},
            "$push": {"payment_history": payment_record}
        }
    )
    
    return {"message": "Pago registrado", "new_status": new_status}

@api_router.delete("/accounts-receivable/{account_id}")
async def delete_account_receivable(
    account_id: str,
    user: dict = Depends(get_current_user)
):
    result = await db.accounts_receivable.delete_one({"id": account_id, "user_id": user["id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Cuenta no encontrada")
    return {"message": "Eliminado"}

# ================= TRAVEL GOALS ENDPOINTS =================

@api_router.get("/travel-goals")
async def get_travel_goals(user: dict = Depends(get_current_user)):
    """Get all travel goals"""
    goals = await db.travel_goals.find({"user_id": user["id"]}).to_list(50)
    
    today = datetime.now(timezone.utc)
    result = []
    for goal in goals:
        goal_data = {k: v for k, v in goal.items() if k != "_id"}
        # Calculate days remaining
        try:
            target = datetime.fromisoformat(goal["target_date"].replace("Z", "+00:00"))
            goal_data["days_remaining"] = max(0, (target - today).days)
        except:
            goal_data["days_remaining"] = 0
        # Calculate progress
        goal_data["progress_percent"] = min(100, round((goal.get("saved_amount", 0) / goal["target_amount"]) * 100))
        # Calculate monthly needed
        if goal_data["days_remaining"] > 0:
            remaining = goal["target_amount"] - goal.get("saved_amount", 0)
            months_remaining = max(1, goal_data["days_remaining"] / 30)
            goal_data["monthly_needed"] = round(remaining / months_remaining, 2)
        else:
            goal_data["monthly_needed"] = 0
        result.append(goal_data)
    
    return {"goals": result, "count": len(result)}

@api_router.post("/travel-goals")
async def create_travel_goal(
    goal: TravelGoalCreate,
    user: dict = Depends(get_current_user)
):
    """Create a new travel goal"""
    doc = {
        "id": str(uuid.uuid4()),
        "user_id": user["id"],
        **goal.model_dump(),
        "saved_amount": 0,
        "status": "active",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.travel_goals.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}

@api_router.put("/travel-goals/{goal_id}")
async def update_travel_goal(
    goal_id: str,
    updates: dict,
    user: dict = Depends(get_current_user)
):
    """Update a travel goal"""
    result = await db.travel_goals.update_one(
        {"id": goal_id, "user_id": user["id"]},
        {"$set": updates}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Meta no encontrada")
    return {"message": "Actualizado"}

@api_router.put("/travel-goals/{goal_id}/add-savings")
async def add_travel_savings(
    goal_id: str,
    amount: float,
    user: dict = Depends(get_current_user)
):
    """Add savings to a travel goal"""
    goal = await db.travel_goals.find_one({"id": goal_id, "user_id": user["id"]})
    if not goal:
        raise HTTPException(status_code=404, detail="Meta no encontrada")
    
    new_saved = goal.get("saved_amount", 0) + amount
    new_status = "completed" if new_saved >= goal["target_amount"] else "active"
    
    await db.travel_goals.update_one(
        {"id": goal_id},
        {"$set": {"saved_amount": new_saved, "status": new_status}}
    )
    
    return {"message": "Ahorro agregado", "new_saved": new_saved, "status": new_status}

@api_router.delete("/travel-goals/{goal_id}")
async def delete_travel_goal(
    goal_id: str,
    user: dict = Depends(get_current_user)
):
    result = await db.travel_goals.delete_one({"id": goal_id, "user_id": user["id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Meta no encontrada")
    return {"message": "Eliminado"}

# ================= CASH FLOW PROJECTION =================

@api_router.get("/cashflow/projection")
async def get_cashflow_projection(user: dict = Depends(get_current_user)):
    """Get 30-day cash flow projection"""
    today = datetime.now(timezone.utc)
    thirty_days = today + timedelta(days=30)
    today_str = today.strftime("%Y-%m-%d")
    thirty_str = thirty_days.strftime("%Y-%m-%d")
    
    # Get expected income
    expected_income = await db.expected_income.find({
        "user_id": user["id"],
        "status": "pending",
        "expected_date": {"$lte": thirty_str}
    }).to_list(100)
    total_expected_income = sum(i.get("amount", 0) for i in expected_income)
    
    # Get accounts receivable
    receivables = await db.accounts_receivable.find({
        "user_id": user["id"],
        "status": {"$in": ["pending", "overdue", "partial"]},
        "due_date": {"$lte": thirty_str}
    }).to_list(100)
    total_receivables = sum(r.get("amount", 0) - r.get("amount_paid", 0) for r in receivables)
    
    # Get scheduled payments
    scheduled = await db.scheduled_payments.find({"user_id": user["id"]}).to_list(100)
    total_scheduled = sum(s.get("amount", 0) for s in scheduled)
    
    # Get credit card minimums
    cards = await db.credit_cards.find({"user_id": user["id"]}).to_list(10)
    total_card_minimums = sum(c.get("minimum_payment", 0) for c in cards)
    
    # Get deferred payments
    deferred = await db.deferred_payments.find({
        "user_id": user["id"],
        "remaining_installments": {"$gt": 0}
    }).to_list(50)
    total_deferred = sum(d.get("monthly_payment", 0) for d in deferred)
    
    # Calculate projection
    total_outflow = total_scheduled + total_card_minimums + total_deferred
    total_inflow = total_expected_income + total_receivables
    projected_balance = total_inflow - total_outflow
    
    # Determine status
    if projected_balance < 0:
        status = "critical"
        message = f"⚠️ Déficit proyectado de ${abs(projected_balance):,.2f} en 30 días"
    elif projected_balance < total_outflow * 0.2:
        status = "warning"
        message = "⚡ Flujo ajustado - considera postergar gastos no esenciales"
    else:
        status = "healthy"
        message = "✅ Flujo de caja saludable para los próximos 30 días"
    
    return {
        "projection": {
            "expected_income": total_expected_income,
            "receivables": total_receivables,
            "total_inflow": total_inflow,
            "scheduled_payments": total_scheduled,
            "card_minimums": total_card_minimums,
            "deferred_payments": total_deferred,
            "total_outflow": total_outflow,
            "projected_balance": projected_balance
        },
        "status": status,
        "message": message,
        "details": {
            "expected_income_count": len(expected_income),
            "receivables_count": len(receivables),
            "scheduled_count": len(scheduled),
            "cards_count": len(cards),
            "deferred_count": len(deferred)
        }
    }

@api_router.get("/reminders")
async def get_reminders(user: dict = Depends(get_current_user)):
    """Get smart reminders for dashboard"""
    today = datetime.now(timezone.utc)
    reminders = []
    
    # 1. Scheduled payments due soon
    scheduled = await db.scheduled_payments.find({"user_id": user["id"]}, {"_id": 0}).to_list(100)
    for payment in scheduled:
        due_day = payment.get("due_day", 1)
        current_month = today.month
        current_year = today.year
        
        if today.day > due_day:
            next_month = current_month + 1 if current_month < 12 else 1
            next_year = current_year if current_month < 12 else current_year + 1
        else:
            next_month = current_month
            next_year = current_year
        
        try:
            due_date = datetime(next_year, next_month, min(due_day, 28))
            days_until = (due_date - today.replace(tzinfo=None)).days
            
            if days_until <= payment.get("reminder_days_before", 2) and days_until >= 0:
                reminders.append({
                    "type": "payment_due",
                    "priority": "high" if days_until == 0 else "medium",
                    "title": f"Pago de {payment['name']}",
                    "message": f"Vence {'hoy' if days_until == 0 else f'en {days_until} días'} - ${payment['amount']:.2f}",
                    "action": f"Pagar con {payment.get('payment_method', 'transferencia')}",
                    "category": payment.get("category"),
                    "amount": payment.get("amount")
                })
        except:
            pass
    
    # 2. Credit card payment reminders
    cards = await db.credit_cards.find({"user_id": user["id"]}, {"_id": 0}).to_list(100)
    for card in cards:
        if card.get("current_balance", 0) > 0:
            payment_day = card.get("payment_due_day", 15)
            current_month = today.month
            current_year = today.year
            
            if today.day > payment_day:
                next_month = current_month + 1 if current_month < 12 else 1
                next_year = current_year if current_month < 12 else current_year + 1
            else:
                next_month = current_month
                next_year = current_year
            
            try:
                due_date = datetime(next_year, next_month, min(payment_day, 28))
                days_until = (due_date - today.replace(tzinfo=None)).days
                
                if days_until <= 5 and days_until >= 0:
                    reminders.append({
                        "type": "card_payment",
                        "priority": "high" if days_until <= 2 else "medium",
                        "title": f"Pago tarjeta {card['name']}",
                        "message": f"Mínimo: ${card.get('minimum_payment', 0):.2f} | Total: ${card['current_balance']:.2f}",
                        "action": f"Vence {'hoy' if days_until == 0 else f'en {days_until} días'}",
                        "card_id": card.get("id"),
                        "apr": card.get("apr")
                    })
            except:
                pass
    
    # 3. Subscription review reminders (every 3 months)
    recurring_expenses = await db.transactions.find({
        "user_id": user["id"],
        "is_recurring": True,
        "transaction_type": "expense"
    }, {"_id": 0}).to_list(100)
    
    # Check for recurring expenses that might need review
    for expense in recurring_expenses[:3]:  # Limit to 3
        reminders.append({
            "type": "subscription_review",
            "priority": "low",
            "title": f"¿Sigues usando {expense.get('description', 'este servicio')}?",
            "message": f"${expense.get('amount', 0):.2f}/mes - Revisa si lo necesitas",
            "action": "Revisar"
        })
    
    # 4. Medical expense reminder (if category is salud)
    medical_expenses = await db.transactions.find({
        "user_id": user["id"],
        "category": "salud",
        "date": {"$gte": (today - timedelta(days=30)).strftime("%Y-%m-%d")}
    }, {"_id": 0}).to_list(10)
    
    for med in medical_expenses[:2]:
        if not med.get("sent_to_insurance"):
            reminders.append({
                "type": "insurance_reminder",
                "priority": "medium",
                "title": "Enviar factura al seguro",
                "message": f"{med.get('description', 'Gasto médico')} - ${med.get('amount', 0):.2f}",
                "action": "Enviar para deducible de prima"
            })
    
    # 5. Motivational message if no urgent reminders
    if not any(r["priority"] == "high" for r in reminders):
        import random
        motivational = [
            "💪 ¡Vas bien! Sigue controlando tus gastos.",
            "🎯 Recuerda tu meta: Gastos fijos 55-65%",
            "💡 Tip: Revisa tus suscripciones mensualmente",
            "🌟 ¡Excelente! No tienes pagos urgentes pendientes",
            "📈 Cada día sin deuda nueva es un paso adelante"
        ]
        reminders.append({
            "type": "motivation",
            "priority": "low",
            "title": random.choice(motivational),
            "message": "",
            "action": None
        })
    
    # Sort by priority
    priority_order = {"high": 0, "medium": 1, "low": 2}
    reminders.sort(key=lambda x: priority_order.get(x.get("priority", "low"), 2))
    
    return reminders

# ================= AI CHATBOT =================

class ChatMessage(BaseModel):
    message: str
    session_id: Optional[str] = None

class ChatResponse(BaseModel):
    response: str
    session_id: str

@api_router.post("/chat", response_model=ChatResponse)
async def chat_with_ai(
    chat_message: ChatMessage,
    user: dict = Depends(get_current_user)
):
    """
    Chat with AI assistant about financial data.
    The AI has context about your transactions, budget, and financial situation.
    """
    if not EMERGENT_LLM_KEY:
        raise HTTPException(status_code=500, detail="API key no configurada")
    
    session_id = chat_message.session_id or f"chat_{user['id']}_{uuid.uuid4().hex[:8]}"
    
    try:
        # Gather user's financial context
        now = datetime.now(timezone.utc)
        start_of_month = now.replace(day=1).strftime("%Y-%m-%d")
        start_of_year = f"{now.year}-01-01"
        
        # Get recent transactions
        recent_transactions = await db.transactions.find(
            {"user_id": user["id"], "date": {"$gte": start_of_month}},
            {"_id": 0, "description": 1, "amount": 1, "category": 1, "date": 1}
        ).sort("date", -1).to_list(50)
        
        # Get monthly totals
        monthly_totals = {}
        for t in recent_transactions:
            cat = t.get("category", "otros")
            monthly_totals[cat] = monthly_totals.get(cat, 0) + t.get("amount", 0)
        
        # Get income summary
        incomes = await db.incomes.find(
            {"user_id": user["id"], "date": {"$gte": start_of_year}},
            {"_id": 0}
        ).to_list(100)
        total_income = sum(i.get("amount", 0) for i in incomes)
        
        # Get credit cards
        cards = await db.credit_cards.find({"user_id": user["id"]}, {"_id": 0}).to_list(10)
        total_debt = sum(c.get("current_balance", 0) for c in cards)
        
        # Build context for AI
        financial_context = f"""
CONTEXTO FINANCIERO DEL USUARIO:

**Presupuesto Mensual (del Excel del usuario):**
- Ingresos esperados: $12,500/mes ($150,000/año)
  - Personal: $7,250/mes
  - APX: $2,500/mes  
  - USA: $2,750/mes

**Categorías de Presupuesto Personal:**
{json.dumps({k: {"nombre": v["name"], "presupuesto_mensual": v.get("monthly_budget", 0)} for k, v in BUDGET_CATEGORIES.items()}, indent=2, ensure_ascii=False)}

**Resumen del mes actual ({now.strftime('%B %Y')}):**
- Gastos por categoría: {json.dumps(monthly_totals, ensure_ascii=False)}
- Total gastado este mes: ${sum(monthly_totals.values()):,.2f}

**Ingresos registrados este año:** ${total_income:,.2f}

**Tarjetas de crédito:**
- Total deuda actual: ${total_debt:,.2f}
- Tarjetas: {', '.join([f"{c['name']} (${c['current_balance']:,.2f})" for c in cards]) if cards else 'Sin tarjetas registradas'}

**Últimas 10 transacciones:**
{chr(10).join([f"- {t['date']}: {t['description']} - ${t['amount']:,.2f} ({t['category']})" for t in recent_transactions[:10]])}

**Metas financieras:**
- Gastos fijos: 55-65% del ingreso
- Ahorro: 10% ($1,250/mes)
- Inversión: 15% ($1,875/mes)
"""

        # Create chat with context
        chat = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=session_id,
            system_message=f"""Eres un asistente financiero personal experto en finanzas familiares y leyes tributarias de Ecuador.
            
Tu rol es ayudar al usuario a:
1. Analizar sus gastos y transacciones
2. Dar consejos para optimizar su presupuesto
3. Responder preguntas sobre su situación financiera
4. Sugerir formas de aumentar ahorros y reducir deudas
5. Explicar temas tributarios del SRI Ecuador

IMPORTANTE:
- Responde SIEMPRE en español
- Sé conciso y práctico
- Usa los datos reales del usuario cuando sea relevante
- Si no sabes algo específico, pide más detalles
- No inventes datos que no tienes

{financial_context}
"""
        ).with_model("openai", "gpt-5.2")
        
        response = await chat.send_message(UserMessage(text=chat_message.message))
        
        # Store conversation in database for history
        await db.chat_history.insert_one({
            "session_id": session_id,
            "user_id": user["id"],
            "user_message": chat_message.message,
            "ai_response": response,
            "timestamp": datetime.now(timezone.utc).isoformat()
        })
        
        return ChatResponse(response=response, session_id=session_id)
        
    except Exception as e:
        logger.error(f"Chat error: {e}")
        raise HTTPException(status_code=500, detail=f"Error en el chat: {str(e)}")

@api_router.get("/chat/history")
async def get_chat_history(
    session_id: Optional[str] = None,
    limit: int = 20,
    user: dict = Depends(get_current_user)
):
    """Get chat history for the user"""
    query = {"user_id": user["id"]}
    if session_id:
        query["session_id"] = session_id
    
    history = await db.chat_history.find(
        query,
        {"_id": 0}
    ).sort("timestamp", -1).limit(limit).to_list(limit)
    
    return {"history": list(reversed(history)), "count": len(history)}

# ================= HEALTH CHECK =================

@api_router.get("/")
async def root():
    return {"message": "FamilyFinance Ecuador API", "version": "1.0.0"}

@api_router.get("/health")
async def health():
    return {"status": "healthy"}

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
async def startup_db_seed():
    """Inicializa los datos base al arrancar la aplicación"""
    logger.info("🚀 Iniciando aplicación - Ejecutando seed de datos...")
    await seed_database(mongo_url, db_name)
    logger.info("✅ Seed de datos completado")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
